# Python 3.7

import sys
import os
import shutil
import time
import datetime
import requests
import json
import traceback
import time
import pymysql, pymysql.cursors
import xlsxwriter as xlsxwriter
from openpyxl import load_workbook

from _env import (
    DB_HOST,
    DB_USER,
    DB_PASS,
    DB_PORT,
    DB_NAME,
    API_URL,
    USERNAME,
    PASSWORD,
    PR_RESULT_DIR as RESULT_DIR,
    PR_SRC_DIR as SRC_DIR,
    PR_SRC_INPROGRESS_DIR as SRC_INPROGRESS_DIR,
    PR_SRC_ACHIEVE_DIR as SRC_ACHIEVE_DIR,
)
from _options_lib import get_option, set_option


def get_token():
    url = API_URL + "/api-token-auth/"
    data = {"username": USERNAME, "password": PASSWORD}
    response = requests.post(url, params={}, json=data)
    response0 = response.content.decode("utf8")
    data0 = json.loads(response0)

    if "token" in data0:
        print("@101 - Token: ", data0["token"])
        return data0["token"]
    else:
        print("@400 - ", data0["non_field_errors"])
        return None


def _update_file_info(mysqlcon, fname, fpath, note):
    print(f"#800 - {fname}, {note}")
    modified_fpath = fpath.replace("../dme_api/", "")
    cursor = mysqlcon.cursor()
    sql = "UPDATE dme_files \
        SET file_path=%s, note=%s \
        WHERE file_name=%s"
    cursor.execute(sql, (modified_fpath, note, fname))
    mysqlcon.commit()


def _insert_file_info(mysqlcon, fname, fpath, note):
    print(f"801 - {fname}, {note}")
    modified_fpath = fpath.replace("../dme_api/", "")
    cursor = mysqlcon.cursor()

    sql = "DELETE FROM dme_files \
        WHERE file_name like %s and file_type=%s"
    cursor.execute(sql, (f"{fname}%", "pricing-rules"))
    mysqlcon.commit()

    sql = "INSERT INTO dme_files \
        (file_name, file_path, z_createdByAccount, note, file_type) \
        VALUES (%s, %s, %s, %s, %s)"
    cursor.execute(sql, (fname, modified_fpath, "script", note, "pricing-rules"))
    mysqlcon.commit()


def _set_rule_type_of_fp(mysqlcon, freight_provider, rule_type):
    cursor = mysqlcon.cursor()
    sql = "UPDATE dme_files \
        SET rule_type_id=%s \
        WHERE id=%s"
    cursor.execute(sql, (int(rule_type[-2:]), freight_provider["id"]))
    mysqlcon.commit()
    print(
        f'# 111 - Updated Freight Provider({freight_provider["fp_company_name"]}) with Rule Type({rule_type}).'
    )


# Set Null if cell is empty or 0
def set_null(value):
    return value if value else None


def read_xls(file):
    wb = load_workbook(filename=file, data_only=True)

    # Check tabs
    if not (
        "Freight Provider" in wb.sheetnames
        and "Timing" in wb.sheetnames
        and "Vehicle" in wb.sheetnames
        and "Availability" in wb.sheetnames
        and "Cost" in wb.sheetnames
        and "Rules" in wb.sheetnames
        and "Postal Zones" in wb.sheetnames
    ):
        print(
            "#910 - Tabs are required - 'Freight Provider', 'Timing', 'Vehicle', 'Availability', 'Cost', 'Rules', 'Postal Zones'"
        )
        exit(1)

    ws = wb["Freight Provider"]
    row = 3
    freight_providers = []
    while True:
        if ws["A%i" % row].value == None:
            break

        freight_provider = {}
        freight_provider["id"] = ws["A%i" % row].value
        freight_provider["fp_company_name"] = ws["B%i" % row].value
        freight_provider["fp_address_country"] = ws["C%i" % row].value
        freight_provider["fp_inactive_date"] = ws["D%i" % row].value
        freight_provider["fp_markupfuel_levy_percent"] = ws["E%i" % row].value
        freight_providers.append(freight_provider)
        row += 1

    ws = wb["Timing"]
    row = 2
    timings = []
    while True:
        if ws["A%i" % row].value == None:
            break

        timing = {}
        timing["id"] = ws["A%i" % row].value
        timing["fp_delivery_service_code"] = ws["B%i" % row].value
        timing["time_uom"] = ws["C%i" % row].value
        timing["min"] = set_null(ws["D%i" % row].value)
        timing["max"] = set_null(ws["E%i" % row].value)

        if timing["max"]:
            timing["fp_03_delivery_hours"] = int(timing["max"]) * 24
        else:
            timing["fp_03_delivery_hours"] = int(timing["min"]) * 24

        timings.append(timing)
        row += 1

    ws = wb["Vehicle"]
    row = 2
    vehicles = []
    while True:
        if ws["A%i" % row].value == None:
            break

        vehicle = {}
        vehicle["id"] = ws["A%i" % row].value
        vehicle["freight_provider_id"] = set_null(ws["B%i" % row].value)
        vehicle["description"] = set_null(ws["C%i" % row].value)
        vehicle["dim_UOM"] = set_null(ws["D%i" % row].value)
        vehicle["max_length"] = set_null(ws["E%i" % row].value)
        vehicle["max_width"] = set_null(ws["F%i" % row].value)
        vehicle["max_height"] = set_null(ws["G%i" % row].value)
        vehicle["mass_UOM"] = set_null(ws["H%i" % row].value)
        vehicle["pallets"] = ws["I%i" % row].value
        vehicle["pallet_UOM"] = set_null(ws["J%i" % row].value)
        vehicle["max_pallet_length"] = ws["K%i" % row].value
        vehicle["max_pallet_width"] = ws["L%i" % row].value
        vehicle["max_pallet_height"] = ws["M%i" % row].value
        vehicle["base_charge"] = set_null(ws["N%i" % row].value)
        vehicle["min_charge"] = set_null(ws["O%i" % row].value)
        vehicle["limited_state"] = set_null(ws["P%i" % row].value)
        vehicle["max_mass"] = set_null(ws["Q%i" % row].value)
        vehicles.append(vehicle)
        row += 1

    ws = wb["Availability"]
    row = 2
    availabilities = []
    while True:
        if ws["A%i" % row].value == None:
            break

        availability = {}
        availability["id"] = ws["A%i" % row].value
        availability["freight_provider_id"] = ws["B%i" % row].value
        availability["code"] = ws["C%i" % row].value
        availability["mon_start"] = set_null(ws["D%i" % row].value)
        availability["mon_end"] = set_null(ws["E%i" % row].value)
        availability["tue_start"] = set_null(ws["F%i" % row].value)
        availability["tue_end"] = set_null(ws["G%i" % row].value)
        availability["wed_start"] = set_null(ws["H%i" % row].value)
        availability["wed_end"] = set_null(ws["I%i" % row].value)
        availability["thu_start"] = set_null(ws["J%i" % row].value)
        availability["thu_end"] = set_null(ws["K%i" % row].value)
        availability["fri_start"] = set_null(ws["L%i" % row].value)
        availability["fri_end"] = set_null(ws["M%i" % row].value)
        availability["sat_start"] = set_null(ws["N%i" % row].value)
        availability["sat_end"] = set_null(ws["O%i" % row].value)
        availability["sun_start"] = set_null(ws["P%i" % row].value)
        availability["sun_end"] = set_null(ws["Q%i" % row].value)
        availabilities.append(availability)
        row += 1

    ws = wb["Cost"]
    row = 2
    costs = []
    while True:
        if ws["A%i" % row].value == None:
            break

        cost = {}
        cost["id"] = ws["A%i" % row].value
        cost["UOM_charge"] = ws["B%i" % row].value
        cost["start_qty"] = set_null(ws["C%i" % row].value)
        cost["end_qty"] = set_null(ws["D%i" % row].value)
        cost["basic_charge"] = set_null(ws["E%i" % row].value)
        cost["min_charge"] = set_null(ws["F%i" % row].value)
        cost["per_UOM_charge"] = set_null(ws["G%i" % row].value)
        cost["oversize_premium"] = set_null(ws["H%i" % row].value)
        cost["oversize_price"] = set_null(ws["I%i" % row].value)
        cost["m3_to_kg_factor"] = set_null(ws["J%i" % row].value)
        cost["dim_UOM"] = set_null(ws["K%i" % row].value)
        cost["price_up_to_length"] = set_null(ws["L%i" % row].value)
        cost["price_up_to_width"] = set_null(ws["M%i" % row].value)
        cost["price_up_to_height"] = set_null(ws["N%i" % row].value)
        cost["weight_UOM"] = set_null(ws["O%i" % row].value)
        cost["price_up_to_weight"] = set_null(ws["P%i" % row].value)
        cost["max_length"] = set_null(ws["Q%i" % row].value)
        cost["max_width"] = set_null(ws["R%i" % row].value)
        cost["max_height"] = set_null(ws["S%i" % row].value)
        cost["max_weight"] = set_null(ws["T%i" % row].value)
        costs.append(cost)
        row += 1

    ws = wb["Rules"]
    row = 2
    rules = []
    while True:
        if ws["A%i" % row].value == None:
            break

        rule = {}
        rule["id"] = ws["A%i" % row].value
        rule["freight_provider_id"] = set_null(ws["B%i" % row].value)
        rule["service_type"] = set_null(ws["C%i" % row].value)
        rule["service_timing_code"] = set_null(ws["D%i" % row].value)
        rule["cost_id"] = set_null(ws["E%i" % row].value)
        rule["de_zone"] = set_null(ws["F%i" % row].value)
        rule["pu_zone"] = set_null(ws["G%i" % row].value)
        rule["timing_id"] = set_null(ws["H%i" % row].value)
        rule["de_postal_code"] = set_null(ws["I%i" % row].value)
        rule["de_state"] = set_null(ws["J%i" % row].value)
        rule["de_suburb"] = set_null(ws["K%i" % row].value)
        rule["pu_postal_code"] = set_null(ws["L%i" % row].value)
        rule["pu_state"] = set_null(ws["M%i" % row].value)
        rule["pu_suburb"] = set_null(ws["N%i" % row].value)
        rule["both_way"] = ws["O%i" % row].value
        rule["vehicle_type"] = set_null(ws["P%i" % row].value)

        if not rule["freight_provider_id"] or not rule["cost_id"]:
            print(f'#409 - Error: Rule({rule["id"]}) missed foreign key(s)')
            exit(1)

        rules.append(rule)
        row += 1

    return freight_providers, timings, vehicles, availabilities, costs, rules


def _populate_vehicle_id(rules, vehicles):
    for rule in rules:
        if not rule["vehicle_type"]:
            rule["vehicle_id"] = None
            del rule["vehicle_type"]
            continue

        for vehicle in vehicles:
            if rule["vehicle_type"] == vehicle["description"]:
                rule["vehicle_id"] = vehicle["id"]
                del rule["vehicle_type"]
                break


def _populate_etd_id(rules, freight_provider, rule_type, timings=[]):
    cursor = mysqlcon.cursor()
    sql = "SELECT * \
        FROM fp_service_etds \
        WHERE freight_provider_id=%s"
    cursor.execute(sql, (freight_provider["id"]))
    etds = cursor.fetchall()

    for rule in rules:
        for etd in etds:
            if rule_type == "rule_type_01":
                if (
                    rule["service_timing_code"].lower()
                    == etd["fp_delivery_service_code"].lower()
                ):
                    rule["etd_id"] = etd["id"]
                    break
            elif rule_type in ["rule_type_02", "rule_type_03"] and timings:
                if (
                    rule["service_timing_code"].lower()
                    == etd["fp_delivery_service_code"].lower()
                    and timings[int(rule["timing_id"]) - 1]["fp_03_delivery_hours"]
                    == etd["fp_03_delivery_hours"]
                ):
                    rule["etd_id"] = etd["id"]
                    break

        del rule["timing_id"]


def get_or_create_freight_provider(token, rule, freight_providers):
    fp_company_name = None

    for freight_provider in freight_providers:
        if rule["freight_provider_id"] == freight_provider["id"]:
            fp_id = freight_provider["id"]
            fp_company_name = freight_provider["fp_company_name"]
            fp_address_country = freight_provider["fp_address_country"]
            break

    if fp_company_name:
        # print("@211 - FP: ", fp_company_name, fp_address_country)

        url = API_URL + "/fp/add/"
        headers = {"Authorization": f"JWT {token}"}
        data = {
            "fp_company_name": fp_company_name,
            "fp_address_country": fp_address_country,
        }
        response = requests.post(url, params={}, json=data, headers=headers)
        response0 = response.content.decode("utf8")

        try:
            data0 = json.loads(response0)

            if int(fp_id) != int(data0["result"]["id"]):
                print(
                    f"@212 - CREATED - xls index: {fp_id}, result: {data0['result']['id']}"
                )
            # s0 = json.dumps(data0, indent=4, sort_keys=True)  # Just for visual
            # print(f"@213 - Pricing result: {s0}")
            data0["result"]["xls_id"] = fp_id
            return data0["result"]
        except Exception as e:
            traceback.print_exc()
            print("@219 Error - ", str(e))
            exit(1)


def get_or_create_objects(token, objects, name, rules=None):
    results = []
    created_count = 0
    exist_count = 0

    for obj in objects:
        # time.sleep(1) # Give some delay for server performance
        # print("@201 - Each Obj: ", obj)
        url = f"{API_URL}/{name}/add/"
        headers = {"Authorization": f"JWT {token}"}
        response = requests.post(url, params={}, json=obj, headers=headers)
        response0 = response.content.decode("utf8")

        try:
            data0 = json.loads(response0)
            if int(obj["id"]) != int(data0["result"]["id"]):
                print(
                    f"@203 - Diff index - xls index: {obj['id']}, result: {data0['result']['id']}"
                )

                if rules and not name in ["vehicles", "availabilities", "fp-cost"]:
                    for rule in rules:
                        if rule[f"{name[:-1]}_id"] == obj["id"]:
                            rule[f"{name[:-1]}_id"] = int(data0["result"]["id"])

                if rules and name == "fp-cost":
                    for rule in rules:
                        if rule[f"cost_id"] == obj["id"]:
                            rule[f"cost_id"] = int(data0["result"]["id"])

            if data0["isCreated"]:
                created_count += 1
            else:
                exist_count += 1

            # s0 = json.dumps(data0, indent=4, sort_keys=True)  # Just for visual
            # print(f"@203 - Pricing result: {s0}")
            data0["result"]["xls_id"] = obj["id"]
            results.append(data0["result"])
        except Exception as e:
            traceback.print_exc()
            print("@209 Error - ", str(e))
            exit(1)

    print(f"@204 - Created count: {created_count}, Existing count: {exist_count}")
    return results


def do_process(mysqlcon, fpath, fname):
    token = get_token()

    if not token:
        print("# 99 - Can't login with givin credentials")
        exit(1)

    print("# 100 - Reading XLS")
    _update_file_info(
        mysqlcon, fname, SRC_INPROGRESS_DIR + fname, "In progress: 0% --- Reading XLS"
    )
    freight_providers, timings, vehicles, availabilities, costs, rules = read_xls(fpath)

    print("# 110 - Checking Rule Type...")
    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 0% --- Checking Rule Type",
    )
    rule_type = fname.split("__")[1]
    first_rule = rules[0]
    freight_provider = get_or_create_freight_provider(
        token, first_rule, freight_providers
    )

    if not freight_provider["rule_type_code"]:
        _set_rule_type_of_fp(mysqlcon, freight_provider, rule_type)
    if rule_type != freight_provider["rule_type_code"]:
        print(
            f"# 501 Error: Freight Provider({freight_provider['fp_company_name']}) has this Rule Type({freight_provider['rule_type_code']}). But imported Rule({rule['id']}) with wrong Rule Type({rule_type})"
        )
        exit(1)

    print("# 102 - Get or Create Vehicles...")
    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 0% --- Get or Create Vehicles...",
    )
    vehicles = get_or_create_objects(token, vehicles, "vehicles", rules)
    print("# 102 - Get or Create Availabilities...")
    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 0% --- Get or Create Availabilities...",
    )
    availabilities = get_or_create_objects(
        token, availabilities, "availabilities", rules
    )
    print("# 103 - Get or Create Costs...")
    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 30% --- Get or Create Costs...",
    )
    costs = get_or_create_objects(token, costs, "fp-cost", rules)

    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 35% --- Populating Vehicle Id",
    )
    _populate_vehicle_id(rules, vehicles)
    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 40% --- Populating ETD Id",
    )
    _populate_etd_id(rules, freight_provider, rule_type, timings)

    print("# 104 - Get or Create Rules...")
    _update_file_info(
        mysqlcon,
        fname,
        SRC_INPROGRESS_DIR + fname,
        "In progress: 40% --- Get or Create Rules...",
    )
    rules = get_or_create_objects(token, rules, "pricing_rules")


if __name__ == "__main__":
    print("#900 Started %s" % datetime.datetime.now())
    time1 = time.time()

    try:
        mysqlcon = pymysql.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASS,
            db=DB_NAME,
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
        )
    except Exception as e:
        traceback.print_exc()
        print("Mysql DB connection error!", e)
        exit(1)

    if not os.path.exists(SRC_DIR):
        print("#901 src dir does not exist! path:", SRC_DIR)
        exit(1)

    try:
        option = get_option(mysqlcon, "pricing_rules")

        if int(option["option_value"]) == 0:
            print("#905 - `pricing_rules` option is OFF")
        elif option["is_running"]:
            print("#905 - `pricing_rules` script is already RUNNING")
        else:
            print("#906 - `pricing_rules` option is ON")
            set_option(mysqlcon, "pricing_rules", True)

            for fname in os.listdir(SRC_DIR):
                fpath = os.path.join(SRC_DIR, fname)

                if os.path.isfile(fpath) and fname.endswith(".xlsx"):
                    try:
                        shutil.move(SRC_DIR + fname, SRC_INPROGRESS_DIR + fname)
                        _update_file_info(
                            mysqlcon,
                            fname,
                            SRC_INPROGRESS_DIR + fname,
                            "In progress: 0%",
                        )
                        file_path = do_process(
                            mysqlcon, SRC_INPROGRESS_DIR + fname, fname
                        )
                        shutil.move(SRC_INPROGRESS_DIR + fname, SRC_ACHIEVE_DIR + fname)
                        _update_file_info(
                            mysqlcon, fname, SRC_ACHIEVE_DIR + fname, "Done: 100%"
                        )
                    except Exception as e:
                        traceback.print_exc()
                        _update_file_info(
                            mysqlcon,
                            fname,
                            SRC_INPROGRESS_DIR + fname,
                            f"Failed... {str(e)}",
                        )

            set_option(mysqlcon, "pricing_rules", False, time1)
    except OSError as e:
        print("Error:", str(e))
        set_option(mysqlcon, "pricing_rules", False, time1)

    mysqlcon.close()
    print("#999 Finished %s" % datetime.datetime.now())
