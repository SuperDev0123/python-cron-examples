from openpyxl import load_workbook
from collections import OrderedDict
import pymysql, pymysql.cursors
import sys
import os
import errno
import shutil
import datetime
import uuid
import redis

# production = False  # Local
production = True  # Prod

if production:
    UPLOAD_DIR = "/var/www/html/dme_api/media/onedrive"
    IMPORT_DIR = "./xls"
    DB_HOST = "deliverme-db.cgc7xojhvzjl.ap-southeast-2.rds.amazonaws.com"  # New db
    DB_USER = "fmadmin"
    DB_PASS = "oU8pPQxh"
    DB_PORT = 3306
    DB_NAME = "dme_db_dev"  # Dev
    # DB_NAME = 'dme_db_prod'  # Prod
else:
    UPLOAD_DIR = "./dir01"
    IMPORT_DIR = "./dir02"
    DB_HOST = "localhost"
    DB_USER = "root"
    DB_PASS = ""
    DB_PORT = 3306
    DB_NAME = "deliver_me"

redis_host = "localhost"
redis_port = 6379
redis_password = ""


def insert_bookingkey(dbcur, booking):
    sql = "INSERT INTO bok_0_bookingkeys (client_booking_id, filename, success, \
                timestampCreated, v_client_pk_consigment_num, l_000_client_acct_number, \
                l_011_client_warehouse_id, l_012_client_warehouse_name) \
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
    dbcur.execute(sql, [v for _, v in booking.items()])


def insert_header(dbcur, header):
    sql = "INSERT INTO `bok_1_headers` \
               (`client_booking_id`, `b_003_b_service_name`, \
                `b_500_b_client_cust_job_code`, `b_054_b_del_company`, `b_000_b_total_lines`, \
                `b_058_b_del_address_suburb`, `b_057_b_del_address_state`, \
                `b_059_b_del_address_postalcode`, `v_client_pk_consigment_num`, `total_kg`, \
                `fk_client_warehouse_id`, `b_021_b_pu_avail_from_date`, `b_022_b_pu_avail_from_time_hour`, `b_024_b_pu_by_date`, \
                `b_025_b_pu_by_time_hour`, `b_047_b_del_avail_from_date`, `b_048_b_del_avail_from_time_hour`, \
                `b_050_b_del_by_date`, `b_051_b_del_by_time_hour`, `b_client_warehouse_code`, `b_028_b_pu_company`, \
                `b_029_b_pu_address_street_1`, `b_030_b_pu_address_street_2`, `b_031_b_pu_address_state`, \
                `b_033_b_pu_address_postalcode`, `b_032_b_pu_address_suburb`, `b_035_b_pu_contact_full_name`, \
                `b_038_b_pu_phone_main`, `b_037_b_pu_email`, \
                `b_015_b_pu_instructions_contact`, `b_055_b_del_address_street_1`, `b_044_b_del_instructions_address`, \
                `b_043_b_del_instructions_contact`, `b_064_b_del_phone_main`, `b_063_b_del_email`, \
                `b_005_b_created_for`, `b_006_b_created_for_email`, `b_000_1_b_clientReference_RA_Numbers`, \
                `z_test`, `b_client_sales_inv_num`, `b_client_order_num`, \
                `b_client_del_note_num`, `date_processed`, `z_createdTimeStamp`, \
                `success`, \
                `b_007_b_ready_status`, `b_023_b_pu_avail_from_time_minute`, \
                `b_026_b_pu_by_time_minute`, `b_clientPU_Warehouse`, `b_027_b_pu_address_type`, \
                `b_034_b_pu_address_country`, `pu_addressed_saved`, \
                `b_039_b_pu_phone_mobile`, `b_036_b_pu_email_group`, `b_040_b_pu_communicate_via`, \
                `b_014_b_pu_handling_instructions`, `b_016_b_pu_instructions_address`, `b_017_b_pu_warehouse_num`, \
                `b_018_b_pu_warehouse_bay`, `b_049_b_del_avail_from_time_minute`, `b_052_b_del_by_time_minute`, \
                `b_053_b_del_address_type`, `b_056_b_del_address_street_2`, `b_060_b_del_address_country`, \
                `de_to_addressed_saved`, `b_061_b_del_contact_full_name`, `b_065_b_del_phone_mobile`, \
                `b_062_b_del_email_group`, `b_066_b_del_communicate_via`, `b_046_b_del_warehouse_number`, \
                `b_045_b_del_warehouse_bay`, `b_010_b_notes`, `b_008_b_category`, \
                `b_client_max_book_amount`, `b_100_client_price_paid_or_quoted`, `b_009_b_priority`, \
                `b_012_b_driver_bring_connote`, `b_019_b_pu_tail_lift`, `b_020_b_pu_num_operators`, \
                `b_041_b_del_tail_lift`, `b_042_b_del_num_operators`, `b_013_b_package_job`, \
                `b_001_b_freight_provider`, `vx_serviceType_XXX`, `fp_pu_id`, \
                `b_002_b_vehicle_type`, `zb_101_text_1`, `zb_102_text_2`, \
                `zb_103_text_3`, `zb_104_text_4`, `zb_105_text_5`, \
                `zb_121_integer_1`, `zb_122_integer_2`, `zb_123_integer_3`, \
                `zb_124_integer_4`, `zb_125_integer_5`, `zb_131_decimal_1`, \
                `zb_132_decimal_2`, `zb_133_decimal_3`, `zb_134_decimal_4`, \
                `zb_135_decimal_5`, `zb_141_date_1`, `zb_142_date_2`, \
                `zb_143_date_3`, `zb_144_date_4`, `zb_145_date_5`, \
                `pk_header_id`, `fk_client_id`, `b_000_3_consignment_number`) \
           VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    args = [v for _, v in header.items()]
    dbcur.execute(sql, args)


def insert_line(dbcur, line):
    sql = "INSERT INTO bok_2_lines (client_booking_id, l_501_client_UOM, l_009_weight_per_each, \
                l_010_totaldim, l_500_client_run_code, l_003_item, \
                v_client_pk_consigment_num, l_cubic_weight, l_002_qty, \
                l_001_type_of_packaging, l_005_dim_length, l_006_dim_width, \
                l_007_dim_height, `date_processed`, `z_createdTimeStamp`, \
                `success`, \
                `e_pallet_type`, `client_item_number`, `e_item_type`, \
                `client_item_reference`, `l_004_dim_UOM`, `l_008_weight_UOM`, \
                `zbl_101_text_1`, `zbl_102_text_2`, `zbl_103_text_3`, \
                `zbl_104_text_4`, `zbl_105_Text_5`, `zbl_121_integer_1`, \
                `zbl_122_integer_2`, `zbl_123_integer_3`, `zbl_124_integer_4`, \
                `zbl_125_integer_5`, `zbl_131_decimal_1`, `zbl_132_decimal_2`, \
                `zbl_133_decimal_3`, `zbl_134_decimal_4`, `zbl_135_decimal_5`, \
                `zbl_141_date_1`, `zbl_142_date_2`, `zbl_143_date_3`, \
                `zbl_144_date_4`, `zbl_145_date_5`, `pk_booking_lines_id`) \
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s)"
    args = [v for _, v in line.items()]
    dbcur.execute(sql, args)


def insert_line_detail(dbcur, line_detail):
    sql = "INSERT INTO bok_3_lines_data( \
                client_booking_id, ld_002_model_number, ld_003_item_description, \
                ld_001_qty, ld_004_fault_description, ld_007_gap_ra, \
                ld_005_item_serial_number, ld_006_insurance_value, ld_008_client_ref_number, \
                zbld_101_text_1, `zbld_102_text_2`, `zbld_103_text_3`, \
                zbld_104_text_4, `zbld_105_text_5`, `zbld_121_integer_1`, \
                zbld_122_integer_2, `zbld_123_integer_3`, `zbld_124_integer_4`, \
                zbld_125_integer_5, `zbld_131_decimal_1`, `zbld_132_decimal_2`, \
                zbld_133_decimal_3, `zbld_134_decimal_4`, `zbld_135_decimal_5`, \
                zbld_141_date_1, `zbld_142_date_2`, `zbld_143_date_3`, \
                zbld_144_date_4, `zbld_145_date_5`, `fk_booking_lines_id`, \
                success, v_client_pk_consigment_num) \
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s)"
    args = [v for _, v in line_detail.items()]
    dbcur.execute(sql, args)


def insert_log(dbcon, fname, success, date):
    sql = "INSERT INTO Log (__id, filename, success, date) VALUES (0, %s, %s, %s)"
    args = [fname, "1" if success else "0", date.isoformat()]
    with dbcon.cursor() as cur:
        cur.execute(sql, args)


def checkLen(str, limit):
    if str is not None:
        if len(str) > limit:
            return False

    return True


def do_check(dbcon, filename):
    # wb = load_workbook(filename=filename, data_only=True)

    # if 'Interstate and TAS' in wb.sheetnames:
    #     worksheet0 = wb['Interstate and TAS']
    # elif 'ALLIED EXPRESS' in wb.sheetnames:
    #     worksheet0 = wb['ALLIED EXPRESS']

    # row = 6
    check_result = {}
    errors = []
    errorCnt = 0

    # if not checkLen(filename, 128):
    #     error = {}
    #     error['text'] = "File name shouldn't be longer than 128 characters"
    #     error['cell'] = ''
    #     errors.append(error)

    # while True:
    #     headers, lines = OrderedDict(), OrderedDict()
    #     pk_header = worksheet0['H%i' % row].value
    #
    #     if worksheet0['H%i' % row].value is None:
    #         # error = {}
    #         # error['text'] = 'HeaderID is required'
    #         # error['cell'] = 'H' + str(row)
    #         # errors.append(error)
    #         break
    #
    #     if not checkLen(str(worksheet0['H%i' % row].value), 64):
    #         error = {}
    #         error['text'] = "HeaderID field shouldn't be longer than 64 characters"
    #         error['cell'] = 'H' + str(row)
    #         errors.append(error)
    #
    #     if worksheet0['B%i' % row].value is None:
    #         error = {}
    #         error['text'] = "Date Shipped field is required"
    #         error['cell'] = 'B' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(str(worksheet0['F%i' % row].value), 20):
    #         error = {}
    #         error['text'] = "Ship-to field shouldn't be longer than 20 characters"
    #         error['cell'] = 'F' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['G%i' % row].value, 100):
    #         error = {}
    #         error['text'] = "Ship-to field shouldn't be longer than 100 characters"
    #         error['cell'] = 'G' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['P%i' % row].value, 100):
    #         error = {}
    #         error['text'] = "Street field shouldn't be longer than 100 characters"
    #         error['cell'] = 'P' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['Q%i' % row].value, 40):
    #         error = {}
    #         error['text'] = "Suburb field shouldn't be longer than 40 characters"
    #         error['cell'] = 'Q' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['R%i' % row].value, 20):
    #         error = {}
    #         error['text'] = "State field shouldn't be longer than 20 characters"
    #         error['cell'] = 'R' + str(row)
    #         errors.append(error)
    #
    #     if worksheet0['L%i' % row].value is None:
    #         error = {}
    #         error['text'] = "Line field is required"
    #         error['cell'] = 'L' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['L%i' % row].value, 31):
    #         error = {}
    #         error['text'] = "Line field shouldn't be longer than 31 characters"
    #         error['cell'] = 'L' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['T%i' % row].value, 31):
    #         error = {}
    #         error['text'] = "Run Code field shouldn't be longer than 31 characters"
    #         error['cell'] = 'T' + str(row)
    #         errors.append(error)
    #
    #     if not checkLen(worksheet0['K%i' % row].value, 128):
    #         error = {}
    #         error['text'] = "Description field shouldn't be longer than 128 characters"
    #         error['cell'] = 'K' + str(row)
    #         errors.append(error)
    #
    #     row += 1

    if len(errors) > 0:
        check_result["status"] = False
    else:
        check_result["status"] = True

    check_result["errors"] = errors
    return check_result


def do_import(dbcon, redisCon, filename):
    wb = load_workbook(filename=filename, data_only=True)
    data = {}
    datas = []
    warehouse_id = []
    warehouse_name = []
    warehouse_code = []
    warehouse_success_type = []
    import_type = 0

    if "Interstate and TAS" in wb.sheetnames:
        worksheet0 = wb["Interstate and TAS"]
    elif "ALLIED EXPRESS" in wb.sheetnames:
        worksheet0 = wb["ALLIED EXPRESS"]
    elif (
        "Import Headers and Lines" in wb.sheetnames
        and "Import Detail Lines" in wb.sheetnames
    ):
        import_type = 1
        worksheet0 = wb["Import Headers and Lines"]
        worksheet1 = wb["Import Detail Lines"]
    else:
        return

    print("Import type: ", import_type)

    if import_type == 0:
        while True:
            header, line = OrderedDict(), OrderedDict()
            row = 6
            pk_header = worksheet0["H%i" % row].value

            if pk_header == None:  # Best way to determine end of row?
                break

            if not worksheet0["AK%i" % row].value is None:
                warehouse_code.append(worksheet0["AK%i" % row].value)

                with dbcon.cursor() as cur:  # get warehouse name from id
                    sql = "SELECT `name`, `pk_id_client_warehouses`, `success_type` FROM `dme_client_warehouses` WHERE `client_warehouse_code`=%s"
                    cur.execute(sql, warehouse_code[row - 6])
                    result = cur.fetchone()
                    if row == None:
                        warehouse_code.append("")
                        warehouse_name.append("")
                        warehouse_id.append(100)
                        warehouse_success_type.append("")
                    else:
                        warehouse_name.append(result["name"])
                        warehouse_id.append(result["pk_id_client_warehouses"])
                        warehouse_success_type.append(result["success_type"])
            else:
                warehouse_code.append("")
                warehouse_name.append("")
                warehouse_id.append(100)
                warehouse_success_type.append("")

            # Header
            header["client_booking_id"] = pk_header
            # header['b_021_b_pu_avail_from_date'] = ''
            # if not worksheet0['B%i' % row].value is None:
            #     x = worksheet0['B%i' % row].value.split('.')
            #     header['avail_from_date'] = '%s-%s-%s' % (x[2], x[1], x[0])
            header["service_name"] = "R"  # worksheet0['E%i' % row].value
            header["client_cust_job_code"] = worksheet0["F%i" % row].value
            header["company"] = worksheet0["G%i" % row].value
            header["total_lines"] = worksheet0["M%i" % row].value
            header["address_suburb"] = worksheet0["R%i" % row].value
            header["address_state"] = worksheet0["S%i" % row].value
            header["address_postalcode"] = worksheet0["T%i" % row].value
            header["v_client_pk_consigment_num"] = pk_header
            header["total_kg"] = worksheet0["N%i" % row].value
            header["fk_client_warehouse_id"] = warehouse_id[row - 6]
            header["puPickUpAvailFrom_Date"] = worksheet0["AC%i" % row].value
            header["b_022_b_pu_avail_from_time_hour"] = worksheet0["AD%i" % row].value
            header["b_024_b_pu_by_date"] = worksheet0["AE%i" % row].value
            header["b_025_b_pu_by_time_hour"] = worksheet0["AF%i" % row].value
            header["b_047_b_del_avail_from_date"] = worksheet0["AG%i" % row].value
            header["b_048_b_del_avail_from_time_hour"] = worksheet0["AH%i" % row].value
            header["b_050_b_del_by_date"] = worksheet0["AI%i" % row].value
            header["b_051_b_del_by_time_hour"] = worksheet0["AJ%i" % row].value
            header["b_client_warehouse_code"] = worksheet0["AK%i" % row].value
            header["b_028_b_pu_company"] = worksheet0["AL%i" % row].value
            header["b_029_b_pu_address_street_1"] = worksheet0["AM%i" % row].value
            header["b_030_b_pu_address_street_2"] = worksheet0["AN%i" % row].value
            header["b_031_b_pu_address_state"] = worksheet0["AO%i" % row].value
            header["b_033_b_pu_address_postalcode"] = worksheet0["AP%i" % row].value
            header["b_032_b_pu_address_suburb"] = worksheet0["AQ%i" % row].value
            header["b_035_b_pu_contact_full_name"] = worksheet0["AR%i" % row].value
            header["b_038_b_pu_phone_main"] = worksheet0["AS%i" % row].value
            header["b_037_b_pu_email"] = worksheet0["AT%i" % row].value
            header["b_015_b_pu_instructions_contact"] = worksheet0["AU%i" % row].value
            header["b_055_b_del_address_street_1"] = worksheet0["AV%i" % row].value
            header["b_044_b_del_instructions_address"] = worksheet0["AW%i" % row].value
            header["b_043_b_del_instructions_contact"] = worksheet0["AX%i" % row].value
            header["b_064_b_del_phone_main"] = worksheet0["AY%i" % row].value
            header["b_063_b_del_email"] = worksheet0["AZ%i" % row].value
            header["b_005_b_created_for"] = worksheet0["BA%i" % row].value
            header["b_006_b_created_for_email"] = worksheet0["BB%i" % row].value
            header["b_000_1_b_clientReference_RA_Numbers"] = worksheet0[
                "BC%i" % row
            ].value
            header["z_test"] = worksheet0["BD%i" % row].value
            header["b_client_sales_inv_num"] = worksheet0["BE%i" % row].value
            header["b_client_order_num"] = worksheet0["BF%i" % row].value
            header["b_client_del_note_num"] = worksheet0["BG%i" % row].value
            header["date_processed"] = datetime.datetime.now()
            header["z_createdTimeStamp"] = datetime.datetime.now()
            header["success"] = warehouse_success_type[0]
            header["b_007_b_ready_status"] = None
            header["b_023_b_pu_avail_from_time_minute"] = None
            header["b_026_b_pu_by_time_minute"] = None
            header["b_clientPU_Warehouse"] = None
            header["b_027_b_pu_address_type"] = None
            header["b_034_b_pu_address_country"] = None
            header["pu_addressed_saved"] = None
            header["b_039_b_pu_phone_mobile"] = None
            header["b_036_b_pu_email_group"] = None
            header["b_040_b_pu_communicate_via"] = None
            header["b_014_b_pu_handling_instructions"] = None
            header["b_016_b_pu_instructions_address"] = None
            header["b_017_b_pu_warehouse_num"] = None
            header["b_018_b_pu_warehouse_bay"] = None
            header["b_049_b_del_avail_from_time_minute"] = None
            header["b_052_b_del_by_time_minute"] = None
            header["b_053_b_del_address_type"] = None
            header["b_056_b_del_address_street_2"] = None
            header["b_060_b_del_address_country"] = None
            header["de_to_addressed_saved"] = None
            header["b_061_b_del_contact_full_name"] = None
            header["b_065_b_del_phone_mobile"] = None
            header["b_062_b_del_email_group"] = None
            header["b_066_b_del_communicate_via"] = None
            header["b_046_b_del_warehouse_number"] = None
            header["b_045_b_del_warehouse_bay"] = None
            header["b_010_b_notes"] = None
            header["b_008_b_category"] = None
            header["b_client_max_book_amount"] = None
            header["b_100_client_price_paid_or_quoted"] = None
            header["b_009_b_priority"] = None
            header["b_012_b_driver_bring_connote"] = None
            header["b_019_b_pu_tail_lift"] = None
            header["b_020_b_pu_num_operators"] = None
            header["b_041_b_del_tail_lift"] = None
            header["b_042_b_del_num_operators"] = None
            header["b_013_b_package_job"] = None
            header["b_001_b_freight_provider"] = None
            header["vx_serviceType_XXX"] = None
            header["fp_pu_id"] = None
            header["b_002_b_vehicle_type"] = None
            header["zb_101_text_1"] = None
            header["zb_102_text_2"] = None
            header["zb_103_text_3"] = None
            header["zb_104_text_4"] = None
            header["zb_105_Text_5"] = None
            header["zb_121_integer_1"] = None
            header["zb_122_integer_2"] = None
            header["zb_123_integer_3"] = None
            header["zb_124_integer_4"] = None
            header["zb_125_integer_5"] = None
            header["zb_131_decimal_1"] = None
            header["zb_132_decimal_2"] = None
            header["zb_133_decimal_3"] = None
            header["zb_134_decimal_4"] = None
            header["zb_135_decimal_5"] = None
            header["zb_141_date_1"] = None
            header["zb_142_date_2"] = None
            header["zb_143_date_3"] = None
            header["zb_144_date_4"] = None
            header["zb_145_date_5"] = None
            header["pk_header_id"] = pk_header
            header["fk_client_id"] = redisCon.get(
                os.path.basename(filename) + "_l_000_client_acct_number"
            )

            # Line
            line["client_booking_id"] = pk_header
            line["client_UOM"] = worksheet0["L%i" % row].value
            line["weight_per_each"] = worksheet0["N%i" % row].value
            line["totaldim"] = worksheet0["O%i" % row].value
            line["client_run_code"] = worksheet0["T%i" % row].value
            line["l_item"] = worksheet0["K%i" % row].value
            line["v_client_pk_consigment_num"] = pk_header
            line["l_cubic_weight"] = worksheet0["O%i" % row].value
            line["qty"] = worksheet0["M%i" % row].value
            line["l_001_type_of_packaging"] = worksheet0["Y%i" % row].value
            line["l_005_dim_length"] = worksheet0["Z%i" % row].value
            line["l_006_dim_width"] = worksheet0["AA%i" % row].value
            line["l_007_dim_height"] = worksheet0["AB%i" % row].value
            line["date_processed"] = datetime.datetime.now()
            line["z_createdTimeStamp"] = datetime.datetime.now()
            line["success"] = warehouse_success_type[0]
            line["e_pallet_type"] = None
            line["client_item_number"] = None
            line["e_item_type"] = None
            line["client_item_reference"] = None
            line["l_004_dim_UOM"] = None
            line["l_008_weight_UOM"] = None
            line["zbl_101_text_1"] = None
            line["zbl_102_text_2"] = None
            line["zbl_103_text_3"] = None
            line["zbl_104_text_4"] = None
            line["zbl_105_Text_5"] = None
            line["zbl_121_integer_1"] = None
            line["zbl_122_integer_2"] = None
            line["zbl_123_integer_3"] = None
            line["zbl_124_integer_4"] = None
            line["zbl_125_integer_5"] = None
            line["zbl_131_decimal_1"] = None
            line["zbl_132_decimal_2"] = None
            line["zbl_133_decimal_3"] = None
            line["zbl_134_decimal_4"] = None
            line["zbl_135_decimal_5"] = None
            line["zbl_141_date_1"] = None
            line["zbl_142_date_2"] = None
            line["zbl_143_date_3"] = None
            line["zbl_144_date_4"] = None
            line["zbl_145_date_5"] = None

            data[pk_header] = {
                "header": header,
                "line": line,
                "client_booking_id": str(uuid.uuid1()),
            }
            datas.append(data)
            row += 1

        # Insert to DB
        with dbcon.cursor() as cur:
            prev_pk_header_id = ""
            prev_pk_booking_id = ""
            ind = 0

            for data in datas:
                booking = OrderedDict()
                booking["client_booking_id"] = data["client_booking_id"]
                booking["filename"] = os.path.basename(filename)
                booking["success"] = warehouse_success_type[0]
                booking["timestampCreated"] = datetime.datetime.now().isoformat()
                booking["v_client_pk_consigment_num"] = data["header"][
                    "v_client_pk_consigment_num"
                ]
                booking["l_000_client_acct_number"] = redisCon.get(
                    os.path.basename(filename) + "_l_000_client_acct_number"
                )
                booking["l_011_client_warehouse_id"] = warehouse_id[ind]
                booking["l_012_client_warehouse_name"] = warehouse_name[ind]

                if not prev_pk_header_id == data["header"]["pk_header_id"]:
                    prev_pk_header_id = data["header"]["pk_header_id"]
                    prev_pk_booking_id = data["client_booking_id"]
                    data["header"]["client_booking_id"] = prev_pk_booking_id
                    insert_bookingkey(cur, booking)
                    insert_header(cur, data["header"])

                data["line"]["client_booking_id"] = prev_pk_booking_id
                insert_line(cur, line)

                ind += 1
    elif import_type == 1:
        row = 6
        while True:
            header, line = OrderedDict(), OrderedDict()
            pk_header = worksheet0["A%i" % row].value
            if pk_header == None:  # Best way to determine end of row?
                break

            if not worksheet0["I%i" % row].value is None:
                warehouse_code.append(worksheet0["I%i" % row].value)

                with dbcon.cursor() as cur:  # get warehouse name from id
                    sql = "SELECT `name`, `pk_id_client_warehouses`, `success_type` FROM `dme_client_warehouses` WHERE `client_warehouse_code`=%s"
                    cur.execute(sql, warehouse_code[row - 6])
                    result = cur.fetchone()
                    if row == None:
                        warehouse_code.append("")
                        warehouse_name.append("")
                        warehouse_id.append(100)
                        warehouse_success_type.append("")
                    else:
                        warehouse_name.append(result["name"])
                        warehouse_id.append(result["pk_id_client_warehouses"])
                        warehouse_success_type.append(result["success_type"])
            else:
                warehouse_code.append("")
                warehouse_name.append("")
                warehouse_id.append(100)
                warehouse_success_type.append("")

            # Header
            header["client_booking_id"] = pk_header
            header["service_name"] = worksheet0["BW%i" % row].value
            header["client_cust_job_code"] = worksheet0["DI%i" % row].value
            header["company"] = worksheet0["AK%i" % row].value
            header["total_lines"] = 0
            header["address_suburb"] = worksheet0["AO%i" % row].value
            header["address_state"] = worksheet0["AP%i" % row].value
            header["address_postalcode"] = worksheet0["AR%i" % row].value
            header["v_client_pk_consigment_num"] = worksheet0["BV%i" % row].value
            header["total_kg"] = 0
            header["fk_client_warehouse_id"] = warehouse_id[row - 6]
            header["puPickUpAvailFrom_Date"] = worksheet0["C%i" % row].value
            header["b_022_b_pu_avail_from_time_hour"] = worksheet0["D%i" % row].value
            header["b_024_b_pu_by_date"] = worksheet0["F%i" % row].value
            header["b_025_b_pu_by_time_hour"] = worksheet0["G%i" % row].value
            header["b_047_b_del_avail_from_date"] = worksheet0["AE%i" % row].value
            header["b_048_b_del_avail_from_time_hour"] = worksheet0["AF%i" % row].value
            header["b_050_b_del_by_date"] = worksheet0["AH%i" % row].value
            header["b_051_b_del_by_time_hour"] = worksheet0["AI%i" % row].value
            header["b_client_warehouse_code"] = worksheet0["I%i" % row].value
            header["b_028_b_pu_company"] = worksheet0["K%i" % row].value
            header["b_029_b_pu_address_street_1"] = worksheet0["M%i" % row].value
            header["b_030_b_pu_address_street_2"] = worksheet0["N%i" % row].value
            header["b_031_b_pu_address_state"] = worksheet0["P%i" % row].value
            header["b_033_b_pu_address_postalcode"] = worksheet0["R%i" % row].value
            header["b_032_b_pu_address_suburb"] = worksheet0["O%i" % row].value
            header["b_035_b_pu_contact_full_name"] = worksheet0["T%i" % row].value
            header["b_038_b_pu_phone_main"] = worksheet0["U%i" % row].value
            header["b_037_b_pu_email"] = worksheet0["W%i" % row].value
            header["b_015_b_pu_instructions_contact"] = worksheet0["AA%i" % row].value
            header["b_055_b_del_address_street_1"] = worksheet0["AM%i" % row].value
            header["b_044_b_del_instructions_address"] = worksheet0["BA%i" % row].value
            header["b_043_b_del_instructions_contact"] = worksheet0["AZ%i" % row].value
            header["b_064_b_del_phone_main"] = worksheet0["AU%i" % row].value
            header["b_063_b_del_email"] = worksheet0["AW%i" % row].value
            header["b_005_b_created_for"] = worksheet0["BN%i" % row].value
            header["b_006_b_created_for_email"] = ""
            header["b_000_1_b_clientReference_RA_Numbers"] = worksheet0[
                "B%i" % row
            ].value
            header["z_test"] = str(worksheet0["BE%i" % row].value)
            header["b_client_sales_inv_num"] = worksheet0["BF%i" % row].value
            header["b_client_order_num"] = worksheet0["BG%i" % row].value
            header["b_client_del_note_num"] = worksheet0["BH%i" % row].value
            header["date_processed"] = datetime.datetime.now()
            header["z_createdTimeStamp"] = datetime.datetime.now()
            header["success"] = warehouse_success_type[0]
            header["b_007_b_ready_status"] = worksheet0["B%i" % row].value
            header["b_023_b_pu_avail_from_time_minute"] = worksheet0["E%i" % row].value
            header["b_026_b_pu_by_time_minute"] = worksheet0["H%i" % row].value
            header["b_clientPU_Warehouse"] = worksheet0["J%i" % row].value
            header["b_027_b_pu_address_type"] = worksheet0["L%i" % row].value
            header["b_034_b_pu_address_country"] = worksheet0["Q%i" % row].value
            header["pu_addressed_saved"] = worksheet0["S%i" % row].value
            header["b_039_b_pu_phone_mobile"] = worksheet0["V%i" % row].value
            header["b_036_b_pu_email_group"] = worksheet0["X%i" % row].value
            header["b_040_b_pu_communicate_via"] = worksheet0["Y%i" % row].value
            header["b_014_b_pu_handling_instructions"] = worksheet0["Z%i" % row].value
            header["b_016_b_pu_instructions_address"] = worksheet0["AB%i" % row].value
            header["b_017_b_pu_warehouse_num"] = worksheet0["AC%i" % row].value
            header["b_018_b_pu_warehouse_bay"] = worksheet0["AD%i" % row].value
            header["b_049_b_del_avail_from_time_minute"] = worksheet0[
                "AG%i" % row
            ].value
            header["b_052_b_del_by_time_minute"] = worksheet0["AJ%i" % row].value
            header["b_053_b_del_address_type"] = worksheet0["AL%i" % row].value
            header["b_056_b_del_address_street_2"] = worksheet0["AN%i" % row].value
            header["b_060_b_del_address_country"] = worksheet0["AQ%i" % row].value
            header["de_to_addressed_saved"] = worksheet0["AS%i" % row].value
            header["b_061_b_del_contact_full_name"] = worksheet0["AT%i" % row].value
            header["b_065_b_del_phone_mobile"] = worksheet0["AV%i" % row].value
            header["b_062_b_del_email_group"] = worksheet0["AX%i" % row].value
            header["b_066_b_del_communicate_via"] = worksheet0["AY%i" % row].value
            header["b_046_b_del_warehouse_number"] = worksheet0["BB%i" % row].value
            header["b_045_b_del_warehouse_bay"] = worksheet0["BC%i" % row].value
            header["b_010_b_notes"] = worksheet0["BD%i" % row].value
            header["b_008_b_category"] = worksheet0["BJ%i" % row].value
            header["b_client_max_book_amount"] = worksheet0["BK%i" % row].value
            header["b_100_client_price_paid_or_quoted"] = worksheet0["BL%i" % row].value
            header["b_009_b_priority"] = worksheet0["BM%i" % row].value
            header["b_012_b_driver_bring_connote"] = worksheet0["BO%i" % row].value
            header["b_019_b_pu_tail_lift"] = worksheet0["BP%i" % row].value
            header["b_020_b_pu_num_operators"] = worksheet0["BQ%i" % row].value
            header["b_041_b_del_tail_lift"] = worksheet0["BR%i" % row].value
            header["b_042_b_del_num_operators"] = worksheet0["BS%i" % row].value
            header["b_013_b_package_job"] = worksheet0["BT%i" % row].value
            header["b_001_b_freight_provider"] = worksheet0["BU%i" % row].value
            header["vx_serviceType_XXX"] = worksheet0["BX%i" % row].value
            header["fp_pu_id"] = worksheet0["BY%i" % row].value
            header["b_002_b_vehicle_type"] = worksheet0["BZ%i" % row].value
            header["zb_101_text_1"] = worksheet0["CA%i" % row].value
            header["zb_102_text_2"] = worksheet0["CB%i" % row].value
            header["zb_103_text_3"] = worksheet0["CC%i" % row].value
            header["zb_104_text_4"] = worksheet0["CD%i" % row].value
            header["zb_105_Text_5"] = worksheet0["CE%i" % row].value
            header["zb_121_integer_1"] = worksheet0["CF%i" % row].value
            header["zb_122_integer_2"] = worksheet0["CG%i" % row].value
            header["zb_123_integer_3"] = worksheet0["CH%i" % row].value
            header["zb_124_integer_4"] = worksheet0["CI%i" % row].value
            header["zb_125_integer_5"] = worksheet0["CJ%i" % row].value
            header["zb_131_decimal_1"] = worksheet0["CK%i" % row].value
            header["zb_132_decimal_2"] = worksheet0["CL%i" % row].value
            header["zb_133_decimal_3"] = worksheet0["CM%i" % row].value
            header["zb_134_decimal_4"] = worksheet0["CN%i" % row].value
            header["zb_135_decimal_5"] = worksheet0["CO%i" % row].value
            header["zb_141_date_1"] = worksheet0["CP%i" % row].value
            header["zb_142_date_2"] = worksheet0["CQ%i" % row].value
            header["zb_143_date_3"] = worksheet0["CR%i" % row].value
            header["zb_144_date_4"] = worksheet0["CS%i" % row].value
            header["zb_145_date_5"] = worksheet0["CT%i" % row].value
            header["pk_header_id"] = pk_header
            header["fk_client_id"] = redisCon.get(
                os.path.basename(filename) + "_l_000_client_acct_number"
            )
            header["b_000_3_consignment_number"] = worksheet0["BV%i" % row].value

            # Line
            line["client_booking_id"] = pk_header
            line["client_UOM"] = worksheet0["DK%i" % row].value
            line["weight_per_each"] = worksheet0["DH%i" % row].value
            line["totaldim"] = ""
            line["client_run_code"] = worksheet0["DJ%i" % row].value
            line["l_item"] = worksheet0["CZ%i" % row].value
            line["v_client_pk_consigment_num"] = pk_header
            line["l_cubic_weight"] = worksheet0["DL%i" % row].value
            line["qty"] = worksheet0["CX%i" % row].value
            line["l_001_type_of_packaging"] = worksheet0["CW%i" % row].value
            line["l_005_dim_length"] = worksheet0["DD%i" % row].value
            line["l_006_dim_width"] = worksheet0["DE%i" % row].value
            line["l_007_dim_height"] = worksheet0["DF%i" % row].value
            line["date_processed"] = datetime.datetime.now()
            line["z_createdTimeStamp"] = datetime.datetime.now()
            line["success"] = warehouse_success_type[0]
            line["e_pallet_type"] = worksheet0["CV%i" % row].value
            line["client_item_number"] = worksheet0["CY%i" % row].value
            line["e_item_type"] = worksheet0["DA%i" % row].value

            line["client_item_reference"] = (
                line["client_item_number"]
                if warehouse_code[0] == "SWAY - HAN"
                else worksheet0["DB%i" % row].value
            )

            line["l_004_dim_UOM"] = worksheet0["DC%i" % row].value
            line["l_008_weight_UOM"] = worksheet0["DG%i" % row].value
            line["zbl_101_text_1"] = worksheet0["DM%i" % row].value
            line["zbl_102_text_2"] = worksheet0["DN%i" % row].value
            line["zbl_103_text_3"] = worksheet0["DO%i" % row].value
            line["zbl_104_text_4"] = worksheet0["DP%i" % row].value
            line["zbl_105_Text_5"] = worksheet0["DQ%i" % row].value
            line["zbl_121_integer_1"] = worksheet0["DR%i" % row].value
            line["zbl_122_integer_2"] = worksheet0["DS%i" % row].value
            line["zbl_123_integer_3"] = worksheet0["DT%i" % row].value
            line["zbl_124_integer_4"] = worksheet0["DU%i" % row].value
            line["zbl_125_integer_5"] = worksheet0["DV%i" % row].value
            line["zbl_131_decimal_1"] = worksheet0["DW%i" % row].value
            line["zbl_132_decimal_2"] = worksheet0["DX%i" % row].value
            line["zbl_133_decimal_3"] = worksheet0["DY%i" % row].value
            line["zbl_134_decimal_4"] = worksheet0["DZ%i" % row].value
            line["zbl_135_decimal_5"] = worksheet0["EA%i" % row].value
            line["zbl_141_date_1"] = worksheet0["EB%i" % row].value
            line["zbl_142_date_2"] = worksheet0["EC%i" % row].value
            line["zbl_143_date_3"] = worksheet0["ED%i" % row].value
            line["zbl_144_date_4"] = worksheet0["EE%i" % row].value
            line["zbl_145_date_5"] = worksheet0["EF%i" % row].value
            line["pk_booking_lines_id"] = worksheet0["EG%i" % row].value

            data = {
                "header": header,
                "line": line,
                "client_booking_id": str(uuid.uuid1()),
            }
            datas.append(data)
            row += 1

        # Line details
        row = 5
        line_details = []

        while True:
            client_booking_id = worksheet1["A%i" % row].value
            if client_booking_id == None:
                break

            line_detail = OrderedDict()
            line_detail["client_booking_id"] = client_booking_id
            line_detail["ld_002_model_number"] = worksheet1["C%i" % row].value
            line_detail["ld_003_item_description"] = worksheet1["D%i" % row].value
            line_detail["ld_001_qty"] = worksheet1["E%i" % row].value
            line_detail["ld_004_fault_description"] = worksheet1["F%i" % row].value
            line_detail["ld_007_gap_ra"] = worksheet1["G%i" % row].value
            line_detail["ld_005_item_serial_number"] = worksheet1["H%i" % row].value
            line_detail["ld_006_insurance_value"] = worksheet1["I%i" % row].value
            line_detail["ld_008_client_ref_number"] = worksheet1["J%i" % row].value
            line_detail["zbld_101_text_1"] = worksheet1["K%i" % row].value
            line_detail["zbld_102_text_2"] = worksheet1["L%i" % row].value
            line_detail["zbld_103_text_3"] = worksheet1["M%i" % row].value
            line_detail["zbld_104_text_4"] = worksheet1["N%i" % row].value
            line_detail["zbld_105_Text_5"] = worksheet1["O%i" % row].value
            line_detail["zbld_121_integer_1"] = worksheet1["P%i" % row].value
            line_detail["zbld_122_integer_2"] = worksheet1["Q%i" % row].value
            line_detail["zbld_123_integer_3"] = worksheet1["R%i" % row].value
            line_detail["zbld_124_integer_4"] = worksheet1["S%i" % row].value
            line_detail["zbld_125_integer_5"] = worksheet1["T%i" % row].value
            line_detail["zbld_131_decimal_1"] = worksheet1["U%i" % row].value
            line_detail["zbld_132_decimal_2"] = worksheet1["V%i" % row].value
            line_detail["zbld_133_decimal_3"] = worksheet1["W%i" % row].value
            line_detail["zbld_134_decimal_4"] = worksheet1["X%i" % row].value
            line_detail["zbld_135_decimal_5"] = worksheet1["Y%i" % row].value
            line_detail["zbld_141_date_1"] = worksheet1["Z%i" % row].value
            line_detail["zbld_142_date_2"] = worksheet1["AA%i" % row].value
            line_detail["zbld_143_date_3"] = worksheet1["AB%i" % row].value
            line_detail["zbld_144_date_4"] = worksheet1["AC%i" % row].value
            line_detail["zbld_145_date_5"] = worksheet1["AD%i" % row].value
            line_detail["fk_booking_lines_id"] = ""
            line_detail["success"] = warehouse_success_type[0]
            line_detail["v_client_pk_consigment_num"] = worksheet1["A%i" % row].value
            line_details.append(line_detail)
            row += 1

        # Insert to DB
        with dbcon.cursor() as cur:
            # Insert line_details
            for line_detail in line_details:
                for data in datas:
                    if (
                        data["header"]["client_booking_id"]
                        == line_detail["client_booking_id"]
                    ):
                        line_detail["client_booking_id"] = data["client_booking_id"]
                        line_detail["fk_booking_lines_id"] = ""
                        insert_line_detail(cur, line_detail)

            prev_pk_header_id = ""
            prev_pk_booking_id = ""
            ind = 0

            for data in datas:
                booking = OrderedDict()
                booking["client_booking_id"] = data["client_booking_id"]
                booking["filename"] = os.path.basename(filename)
                booking["success"] = warehouse_success_type[0]
                booking["timestampCreated"] = datetime.datetime.now().isoformat()
                booking["v_client_pk_consigment_num"] = data["header"][
                    "client_booking_id"
                ]
                booking["l_000_client_acct_number"] = redisCon.get(
                    os.path.basename(filename) + "_l_000_client_acct_number"
                )
                booking["l_011_client_warehouse_id"] = warehouse_id[ind]
                booking["l_012_client_warehouse_name"] = warehouse_name[ind]

                if not prev_pk_header_id == data["header"]["pk_header_id"]:
                    prev_pk_header_id = data["header"]["pk_header_id"]
                    prev_pk_booking_id = data["client_booking_id"]
                    data["header"]["client_booking_id"] = data["client_booking_id"]
                    insert_bookingkey(cur, booking)
                    insert_header(cur, data["header"])

                data["line"]["client_booking_id"] = prev_pk_booking_id
                pk_booking_lines_id = ""  # str(uuid.uuid1())

                # for line_detail in line_details:
                #     if (
                #         data["line"]["pk_booking_lines_id"]
                #         == line_detail["fk_booking_lines_id"]
                #     ):
                #         line_detail["client_booking_id"] = prev_pk_booking_id
                #         line_detail["fk_booking_lines_id"] = pk_booking_lines_id
                #         insert_line_detail(cur, line_detail)

                data["line"]["pk_booking_lines_id"] = pk_booking_lines_id
                insert_line(cur, data["line"])

                ind += 1


def clearCheckHistory(fname, redisCon):
    errorCnt = redisCon.get(fname)

    if errorCnt is not None:
        for index in range(int(errorCnt)):
            redisCon.delete(fname + str(index))

    redisCon.delete(fname)


if __name__ == "__main__":
    print("Running %s" % datetime.datetime.now())
    if not os.path.isdir(UPLOAD_DIR):
        print('Given argument "%s" is not a directory' % UPLOAD_DIR)
        exit(1)

    # Create working directory if necessary
    dirs = [os.path.join(IMPORT_DIR, "imported"), os.path.join(IMPORT_DIR, "error")]
    for dir in dirs:
        try:
            os.mkdir(dir)
        except OSError as e:
            if e.errno != errno.EEXIST:
                print(str(e))
                exit(1)

    try:
        dbcon = pymysql.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASS,
            db=DB_NAME,
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
        )
    except:
        print("Mysql DB connection error!")
        exit(1)

    try:
        redisCon = redis.StrictRedis(
            host=redis_host,
            port=redis_port,
            password=redis_password,
            charset="utf-8",
            decode_responses=True,
        )
    except:
        print("Redis DB connection error!")
        exit(1)

    try:
        success = False
        for fname in os.listdir(UPLOAD_DIR):
            fpath = os.path.join(UPLOAD_DIR, fname)
            print(fpath, os.path.isfile(fpath))

            if os.path.isfile(fpath) and fname.endswith(".xlsx"):
                try:
                    check_result = do_check(dbcon, fpath)
                    # check_result = {'errors': '', 'status': True} # Test Case

                    clearCheckHistory(fname, redisCon)

                    redisCon.set(fname, str(len(check_result["errors"])))
                    if check_result["status"] == False:
                        for index, (key, value) in enumerate(check_result["errors"]):
                            redisCon.set(
                                fname + str(index),
                                str(
                                    check_result["errors"][index].get("text")
                                    + ": "
                                    + check_result["errors"][index].get("cell")
                                ).encode("utf-8"),
                            )
                            print(
                                "Redis set - ",
                                fname + str(index),
                                ":",
                                str(
                                    check_result["errors"][index].get("text")
                                    + ": "
                                    + check_result["errors"][index].get("cell")
                                ).encode("utf-8"),
                            )

                    if check_result["status"] == True:
                        try:
                            do_import(dbcon, redisCon, fpath)

                            shutil.move(
                                fpath, os.path.join(IMPORT_DIR, "imported", fname)
                            )
                            success = True
                            redisCon.set(fname, str(0))
                            print("Redis set - ", fname, ": 0")
                            print("Import successfully ", fname)
                            pass
                        except OSError as e:
                            print(str(e))
                            success = False
                except OSError as e:
                    print(str(e))
                    success = False

                # insert_log(dbcon, fname, success, datetime.datetime.now())
                dbcon.commit()
    except OSError as e:
        print(str(e))

    dbcon.close()

    print("#999 - Finished %s\n\n\n" % datetime.now())

