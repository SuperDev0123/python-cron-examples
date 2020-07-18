# Python 3.7

import sys
import os
import errno
import shutil
import datetime
import uuid
import requests
import json
import traceback
import pymysql, pymysql.cursors
import xlsxwriter as xlsxwriter
from openpyxl import load_workbook

from _env import (
    DB_HOST,
    DB_USER,
    DB_PASS,
    DB_PORT,
    DB_NAME,
    API_URL,
    PO_RESULT_DIR as RESULT_DIR,
    PO_SRC_DIR as SRC_DIR,
    PO_SRC_INPROGRESS_DIR as SRC_INPROGRESS_DIR,
    PO_SRC_ACHIEVE_DIR as SRC_ACHIEVE_DIR,
)
from _options_lib import get_option, set_option


def replace_null(array):
    for item in array:
        for attr in item:
            if item and item[attr] == None:
                item[attr] = ""

    return array


def read_xls(file):
    wb = load_workbook(filename=file, data_only=True)

    if (
        "Import Headers and Lines" in wb.sheetnames
        and "Import Detail Lines" in wb.sheetnames
    ):
        worksheet0 = wb["Import Headers and Lines"]
    else:
        print("#910 - File format is not supported.")
        return

    bookings = []
    booking_lines = []

    row = 6
    last_pk_booking_id = None
    while True:
        pk_booking_id = worksheet0["A%i" % row].value

        if pk_booking_id == None:  # Best way to determine end of row?
            break
        else:
            pk_booking_id = pk_booking_id + "_pricing_only"

        if not last_pk_booking_id or pk_booking_id != last_pk_booking_id:
            last_pk_booking_id = pk_booking_id
            booking = {
                "pk_booking_id": pk_booking_id,
                "puPickUpAvailFrom_Date": str(
                    datetime.datetime.now() + datetime.timedelta(days=1)
                ),
                "b_clientReference_RA_Numbers": worksheet0["B%i" % row].value,
                "puCompany": worksheet0["K%i" % row].value
                if worksheet0["K%i" % row].value
                else "HARDCODED_00",
                "pu_Contact_F_L_Name": worksheet0["T%i" % row].value
                if worksheet0["T%i" % row].value
                else "HARDCODED_01",
                "pu_Email": worksheet0["W%i" % row].value
                if worksheet0["W%i" % row].value
                else "pu@email.com",
                "pu_Phone_Main": worksheet0["U%i" % row].value
                if worksheet0["U%i" % row].value
                else "419294339",
                "pu_Address_Street_1": worksheet0["M%i" % row].value,
                "pu_Address_street_2": worksheet0["N%i" % row].value,
                "pu_Address_Country": worksheet0["Q%i" % row].value,
                "pu_Address_PostalCode": str(worksheet0["R%i" % row].value),
                "pu_Address_State": worksheet0["P%i" % row].value,
                "pu_Address_Suburb": worksheet0["O%i" % row].value,
                "deToCompanyName": worksheet0["AK%i" % row].value
                if worksheet0["AK%i" % row].value
                else "HARDCODED_10",
                "de_to_Contact_F_LName": worksheet0["AT%i" % row].value
                if worksheet0["AT%i" % row].value
                else "HARDCODED_11",
                "de_Email": worksheet0["AW%i" % row].value
                if worksheet0["AW%i" % row].value
                else "de@email.com",
                "de_to_Phone_Main": worksheet0["AV%i" % row].value
                if worksheet0["AV%i" % row].value
                else "419294339",
                "de_To_Address_Street_1": worksheet0["AM%i" % row].value,
                "de_To_Address_Street_2": worksheet0["AN%i" % row].value,
                "de_To_Address_Country": worksheet0["AQ%i" % row].value,
                "de_To_Address_PostalCode": str(worksheet0["AR%i" % row].value),
                "de_To_Address_State": worksheet0["AP%i" % row].value,
                "de_To_Address_Suburb": worksheet0["AO%i" % row].value,
                "client_warehouse_code": worksheet0["I%i" % row].value,
                "vx_serviceName": "R",
                "b_client_name": "Pricing-Only",
            }
            bookings.append(booking)

        booking_line = {
            "fk_booking_id": pk_booking_id,
            "e_dimWidth": worksheet0["DE%i" % row].value,
            "e_dimHeight": worksheet0["DF%i" % row].value,
            "e_dimLength": worksheet0["DD%i" % row].value,
            "e_dimUOM": worksheet0["DC%i" % row].value,
            "e_weightPerEach": worksheet0["DH%i" % row].value,
            "e_weightUOM": worksheet0["DG%i" % row].value,
            "e_item": worksheet0["CZ%i" % row].value,
            "packagingType": worksheet0["CW%i" % row].value,
            "e_qty": worksheet0["CX%i" % row].value,
        }
        booking_lines.append(booking_line)
        row += 1

    return bookings, booking_lines


def do_pricing(booking, booking_lines):
    url = API_URL + "/fp-api/pricing/"
    data = {"booking_id": "", "booking": booking, "booking_lines": booking_lines}
    print(f"@201 - Retrieving Price... : {booking['pk_booking_id']}")
    response = requests.post(url, params={}, json=data)
    response0 = response.content.decode("utf8")
    data0 = json.loads(response0)
    print(
        f"@202 - {data0['message']}, results: {len(data0['results'] if 'results' in data0 else [])}"
    )
    s0 = json.dumps(data0, indent=4, sort_keys=True)  # Just for visual
    # print(f"@210 - Pricing result: {s0}")
    return data0


def do_process(mysqlcon, fpath, fname):
    # Create RESULT_DIR folder if not exist
    if not os.path.exists(RESULT_DIR):
        os.makedirs(RESULT_DIR)

    file_name_without_ext = fname.split(".")[0]
    file_name = f"{file_name_without_ext}_result_{str(datetime.datetime.now().strftime('%d-%m-%Y %H_%M_%S'))}.xlsx"
    file_path = f"{RESULT_DIR}/{file_name}"

    # Initialize XLSWriter
    workbook = xlsxwriter.Workbook(file_path, {"remove_timezone": True})
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({"bold": 1, "align": "left"})
    date_format = workbook.add_format({"num_format": "dd/mm/yyyy"})
    time_format = workbook.add_format({"num_format": "hh:mm:ss"})
    worksheet.set_column(0, 35, width=30)
    row = 2
    col = 0

    # Add header
    worksheet.write("A1", "pk_booking_id", bold)
    worksheet.write("B1", "puPickUpAvailFrom_Date", bold)
    worksheet.write("C1", "b_clientReference_RA_Numbers", bold)
    worksheet.write("D1", "puCompany", bold)
    worksheet.write("E1", "pu_Contact_F_L_Name", bold)
    worksheet.write("F1", "pu_Email", bold)
    worksheet.write("G1", "pu_Phone_Main", bold)
    worksheet.write("H1", "pu_Address_Street_1", bold)
    worksheet.write("I1", "pu_Address_Street_2", bold)
    worksheet.write("J1", "pu_Address_Country", bold)
    worksheet.write("K1", "pu_Address_PostalCode", bold)
    worksheet.write("L1", "pu_Address_State", bold)
    worksheet.write("M1", "pu_Address_Suburb", bold)
    worksheet.write("N1", "deToCompanyName", bold)
    worksheet.write("O1", "de_to_Contact_F_LName", bold)
    worksheet.write("P1", "de_Email", bold)
    worksheet.write("Q1", "de_to_Phone_Main", bold)
    worksheet.write("R1", "de_To_Address_Street_1", bold)
    worksheet.write("S1", "de_To_Address_Street_2", bold)
    worksheet.write("T1", "de_To_Address_Country", bold)
    worksheet.write("U1", "de_To_Address_PostalCode", bold)
    worksheet.write("V1", "de_To_Address_State", bold)
    worksheet.write("W1", "de_To_Address_Suburb", bold)
    worksheet.write("X1", "client_warehouse_code", bold)
    worksheet.write("Y1", "Freight Provider", bold)
    worksheet.write("Z1", "Service Name", bold)
    worksheet.write("AA1", "FP Price", bold)
    worksheet.write("AB1", "DME Price", bold)
    worksheet.write("AC1", "Tax Type", bold)
    worksheet.write("AD1", "Tax Amount", bold)
    worksheet.write("AE1", "Etd", bold)
    worksheet.write("AF1", "mu_percentage_fuel_levy", bold)

    bookings, booking_lines = read_xls(fpath)
    bookings = replace_null(bookings)
    booking_lines = replace_null(booking_lines)

    for booking in bookings:
        lines = []

        for booking_line in booking_lines:
            if booking["pk_booking_id"] == booking_line["fk_booking_id"]:
                lines.append(booking_line)

        pricing_response = do_pricing(booking, lines)

        if "results" in pricing_response:
            lowest = None

            for pricing in pricing_response["results"]:
                if not lowest:
                    lowest = pricing
                else:
                    if int(lowest["client_mu_1_minimum_values"]) >= int(
                        pricing["client_mu_1_minimum_values"]
                    ):
                        lowest = pricing

            if lowest:
                worksheet.write(row, col + 0, booking["pk_booking_id"])
                worksheet.write(row, col + 1, booking["puPickUpAvailFrom_Date"])
                worksheet.write(row, col + 2, booking["b_clientReference_RA_Numbers"])
                worksheet.write(row, col + 3, booking["puCompany"])
                worksheet.write(row, col + 4, booking["pu_Contact_F_L_Name"])
                worksheet.write(row, col + 5, booking["pu_Email"])
                worksheet.write(row, col + 6, booking["pu_Phone_Main"])
                worksheet.write(row, col + 7, booking["pu_Address_Street_1"])
                worksheet.write(row, col + 8, booking["pu_Address_street_2"])
                worksheet.write(row, col + 9, booking["pu_Address_Country"])
                worksheet.write(row, col + 10, booking["pu_Address_PostalCode"])
                worksheet.write(row, col + 11, booking["pu_Address_State"])
                worksheet.write(row, col + 12, booking["pu_Address_Suburb"])
                worksheet.write(row, col + 13, booking["deToCompanyName"])
                worksheet.write(row, col + 14, booking["de_to_Contact_F_LName"])
                worksheet.write(row, col + 15, booking["de_Email"])
                worksheet.write(row, col + 16, booking["de_to_Phone_Main"])
                worksheet.write(row, col + 17, booking["de_To_Address_Street_1"])
                worksheet.write(row, col + 18, booking["de_To_Address_Street_2"])
                worksheet.write(row, col + 19, booking["de_To_Address_Country"])
                worksheet.write(row, col + 20, booking["de_To_Address_PostalCode"])
                worksheet.write(row, col + 21, booking["de_To_Address_State"])
                worksheet.write(row, col + 22, booking["de_To_Address_Suburb"])
                worksheet.write(row, col + 23, booking["client_warehouse_code"])
                worksheet.write(row, col + 24, pricing["fk_freight_provider_id"])
                worksheet.write(row, col + 25, pricing["service_name"])
                worksheet.write(row, col + 26, pricing["fee"])
                worksheet.write(row, col + 27, pricing["client_mu_1_minimum_values"])
                worksheet.write(row, col + 28, pricing["tax_id_1"])
                worksheet.write(row, col + 29, pricing["tax_value_1"])
                worksheet.write(row, col + 30, pricing["etd"])
                worksheet.write(row, col + 31, pricing["mu_percentage_fuel_levy"])

                if row - 2 != 0:
                    note = f"In progress: {int((row - 2) / len(bookings) * 100)}%"
                    _update_file_info(mysqlcon, fname, fpath, note)

                row += 1

    workbook.close()

    return file_path


def _update_file_info(mysqlcon, fname, fpath, note):
    print(f"#800 - {fname}, {note}")
    modified_fpath = fpath.replace("../dme_api/", "")
    cursor = mysqlcon.cursor()
    sql = "UPDATE dme_files \
        SET file_path=%s, note=%s \
        WHERE file_name=%s"
    cursor.execute(sql, (modified_fpath, note, fname))
    mysqlcon.commit()


def _insert_file_info(mysqlcon, fname, fpath, note):
    print(f"801 - {fname}, {note}")
    modified_fpath = fpath.replace("../dme_api/", "")
    cursor = mysqlcon.cursor()

    sql = "DELETE FROM dme_files \
        WHERE file_name like %s and file_type=%s"
    cursor.execute(sql, (f"{fname}%", "pricing-result"))
    mysqlcon.commit()

    sql = "INSERT INTO dme_files \
        (file_name, file_path, z_createdByAccount, note, file_type) \
        VALUES (%s, %s, %s, %s, %s)"
    cursor.execute(sql, (fname, modified_fpath, "script", note, "pricing-result"))
    mysqlcon.commit()


if __name__ == "__main__":
    print("#900 Started %s" % datetime.datetime.now())

    try:
        mysqlcon = pymysql.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASS,
            db=DB_NAME,
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
        )
    except Exception as e:
        print("Mysql DB connection error!", e)
        exit(1)

    if not os.path.exists(SRC_DIR):
        print("#901 src dir does not exist! path:", SRC_DIR)
        exit(1)

    try:
        option = get_option(mysqlcon, "pricing_only")

        if int(option["option_value"]) == 0:
            print("#905 - `pricing_only` option is OFF")
        elif option["is_running"]:
            print("#905 - `pricing_only` script is already RUNNING")
        else:
            print("#906 - `pricing_only` option is ON")
            set_option(mysqlcon, "pricing_only", True)

            for fname in os.listdir(SRC_DIR):
                fpath = os.path.join(SRC_DIR, fname)

                if os.path.isfile(fpath) and fname.endswith(".xlsx"):
                    try:
                        shutil.move(SRC_DIR + fname, SRC_INPROGRESS_DIR + fname)
                        _update_file_info(
                            mysqlcon,
                            fname,
                            SRC_INPROGRESS_DIR + fname,
                            "In progress: 0%",
                        )
                        file_path = do_process(
                            mysqlcon, SRC_INPROGRESS_DIR + fname, fname
                        )
                        shutil.move(SRC_INPROGRESS_DIR + fname, SRC_ACHIEVE_DIR + fname)
                        _update_file_info(
                            mysqlcon, fname, SRC_ACHIEVE_DIR + fname, "Done: 100%"
                        )
                        file_name = file_path.split("/")[-1]
                        _insert_file_info(mysqlcon, file_name, file_path, "Generated")
                    except Exception as e:
                        traceback.print_exc()
                        _update_file_info(
                            mysqlcon,
                            fname,
                            SRC_INPROGRESS_DIR + fname,
                            f"Stopped... {str(e)}",
                        )

            set_option(mysqlcon, "pricing_only", False)
    except OSError as e:
        set_option(mysqlcon, "pricing_only", False)
        print("Error:", str(e))

    mysqlcon.close()
    print("#999 Finished %s" % datetime.datetime.now())
