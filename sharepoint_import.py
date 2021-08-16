# Python 3.7

import sys
import os
import time
import shutil
import uuid
import traceback
import copy
import pytz
from datetime import date, timedelta, datetime
import pymysql, pymysql.cursors
from openpyxl import load_workbook

from _env import (
    DB_HOST,
    DB_USER,
    DB_PASS,
    DB_PORT,
    DB_NAME,
    SI_DOWNLOAD_DIR as DOWNLOAD_DIR,
    SI_IMPORT_DIR as IMPORT_DIR,
    SI_ERROR_DIR as ERROR_DIR,
    EMAIL_USE_TLS,
    EMAIL_HOST,
    EMAIL_PORT,
    EMAIL_HOST_USER,
    EMAIL_HOST_PASSWORD,
)
from _sharepoint_lib import Office365, Site
from _options_lib import get_option, set_option
from _datetime_lib import convert_to_UTC_tz
from _email_lib import send_email


def insert_dme_file(dbcur, file):
    sql = "INSERT INTO dme_files (file_name, z_createdTimestamp, z_createdByAccount, \
                file_type, file_path, file_extension) \
           VALUES (%s, %s, %s, %s, %s, %s)"
    dbcur.execute(sql, [v for _, v in file.items()])


def save_dme_note(dbcur, id, note):
    with dbcon.cursor() as cur:
        sql = "UPDATE `dme_files` SET `note`=%s WHERE `id`=%s"
        cur.execute(sql, (note, id))
        dbcon.commit()


def save_dme_file_path(dbcur, id, file_path):
    with dbcon.cursor() as cur:
        sql = "UPDATE `dme_files` SET `file_path`=%s WHERE `id`=%s"
        cur.execute(sql, (file_path, id))
        dbcon.commit()


def delete_keys(src):
    dest = copy.deepcopy(src)
    dest.pop("id", None)
    dest.pop("header_id", None)
    dest.pop("line_id", None)
    return dest


def insert_header(dbcur, header):
    sql = "INSERT INTO `bok_1_headers` \
               (`client_booking_id`, `b_003_b_service_name`, \
                `b_500_b_client_cust_job_code`, `b_054_b_del_company`, `b_000_b_total_lines`, \
                `b_058_b_del_address_suburb`, `b_057_b_del_address_state`, \
                `b_059_b_del_address_postalcode`, `v_client_pk_consigment_num`, `total_kg`, \
                `fk_client_warehouse_id`, `b_021_b_pu_avail_from_date`, `b_022_b_pu_avail_from_time_hour`, `b_024_b_pu_by_date`, \
                `b_025_b_pu_by_time_hour`, `b_047_b_del_avail_from_date`, `b_048_b_del_avail_from_time_hour`, \
                `b_050_b_del_by_date`, `b_051_b_del_by_time_hour`, `b_client_warehouse_code`, `b_028_b_pu_company`, \
                `b_029_b_pu_address_street_1`, `b_030_b_pu_address_street_2`, `b_031_b_pu_address_state`, \
                `b_033_b_pu_address_postalcode`, `b_032_b_pu_address_suburb`, `b_035_b_pu_contact_full_name`, \
                `b_038_b_pu_phone_main`, `b_037_b_pu_email`, \
                `b_015_b_pu_instructions_contact`, `b_055_b_del_address_street_1`, `b_044_b_del_instructions_address`, \
                `b_043_b_del_instructions_contact`, `b_064_b_del_phone_main`, `b_063_b_del_email`, \
                `b_005_b_created_for`, `b_006_b_created_for_email`, `b_000_1_b_clientReference_RA_Numbers`, \
                `z_test`, `b_client_sales_inv_num`, `b_client_order_num`, \
                `b_client_del_note_num`, `date_processed`, `z_createdTimeStamp`, \
                `success`, \
                `b_007_b_ready_status`, `b_023_b_pu_avail_from_time_minute`, \
                `b_026_b_pu_by_time_minute`, `b_clientPU_Warehouse`, `b_027_b_pu_address_type`, \
                `b_034_b_pu_address_country`, `pu_addressed_saved`, \
                `b_039_b_pu_phone_mobile`, `b_036_b_pu_email_group`, `b_040_b_pu_communicate_via`, \
                `b_014_b_pu_handling_instructions`, `b_016_b_pu_instructions_address`, `b_017_b_pu_warehouse_num`, \
                `b_018_b_pu_warehouse_bay`, `b_049_b_del_avail_from_time_minute`, `b_052_b_del_by_time_minute`, \
                `b_053_b_del_address_type`, `b_056_b_del_address_street_2`, `b_060_b_del_address_country`, \
                `de_to_addressed_saved`, `b_061_b_del_contact_full_name`, `b_065_b_del_phone_mobile`, \
                `b_062_b_del_email_group`, `b_066_b_del_communicate_via`, `b_046_b_del_warehouse_number`, \
                `b_045_b_del_warehouse_bay`, `b_010_b_notes`, `b_008_b_category`, \
                `b_client_max_book_amount`, `b_100_client_price_paid_or_quoted`, `b_009_b_priority`, \
                `b_012_b_driver_bring_connote`, `b_019_b_pu_tail_lift`, `b_020_b_pu_num_operators`, \
                `b_041_b_del_tail_lift`, `b_042_b_del_num_operators`, `b_013_b_package_job`, \
                `b_001_b_freight_provider`, `vx_serviceType_XXX`, `fp_pu_id`, \
                `b_002_b_vehicle_type`, `zb_101_text_1`, `zb_102_text_2`, \
                `zb_103_text_3`, `zb_104_text_4`, `zb_105_text_5`, \
                `zb_121_integer_1`, `zb_122_integer_2`, `zb_123_integer_3`, \
                `zb_124_integer_4`, `zb_125_integer_5`, `zb_131_decimal_1`, \
                `zb_132_decimal_2`, `zb_133_decimal_3`, `zb_134_decimal_4`, \
                `zb_135_decimal_5`, `zb_141_date_1`, `zb_142_date_2`, \
                `zb_143_date_3`, `zb_144_date_4`, `zb_145_date_5`, \
                `pk_header_id`, `fk_client_id`, `b_000_3_consignment_number`, \
                `x_booking_Created_With`) \
           VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    args = [v for _, v in header.items()]
    dbcur.execute(sql, args)


def insert_line(dbcur, line):
    sql = "INSERT INTO bok_2_lines (client_booking_id, l_501_client_UOM, l_009_weight_per_each, \
                l_010_totaldim, l_500_client_run_code, l_003_item, \
                v_client_pk_consigment_num, l_cubic_weight, l_002_qty, \
                l_001_type_of_packaging, l_005_dim_length, l_006_dim_width, \
                l_007_dim_height, `date_processed`, `z_createdTimeStamp`, \
                `success`, \
                `e_pallet_type`, `client_item_number`, `e_item_type`, \
                `client_item_reference`, `l_004_dim_UOM`, `l_008_weight_UOM`, \
                `zbl_101_text_1`, `zbl_102_text_2`, `zbl_103_text_3`, \
                `zbl_104_text_4`, `zbl_105_Text_5`, `zbl_121_integer_1`, \
                `zbl_122_integer_2`, `zbl_123_integer_3`, `zbl_124_integer_4`, \
                `zbl_125_integer_5`, `zbl_131_decimal_1`, `zbl_132_decimal_2`, \
                `zbl_133_decimal_3`, `zbl_134_decimal_4`, `zbl_135_decimal_5`, \
                `zbl_141_date_1`, `zbl_142_date_2`, `zbl_143_date_3`, \
                `zbl_144_date_4`, `zbl_145_date_5`, `pk_booking_lines_id`) \
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s)"
    args = [v for _, v in line.items()]
    dbcur.execute(sql, args)


def insert_line_detail(dbcur, line_detail):
    sql = "INSERT INTO bok_3_lines_data( \
                client_booking_id, ld_002_model_number, ld_003_item_description, \
                ld_001_qty, ld_004_fault_description, ld_007_gap_ra, \
                ld_005_item_serial_number, ld_006_insurance_value, ld_008_client_ref_number, \
                zbld_101_text_1, `zbld_102_text_2`, `zbld_103_text_3`, \
                zbld_104_text_4, `zbld_105_text_5`, `zbld_121_integer_1`, \
                zbld_122_integer_2, `zbld_123_integer_3`, `zbld_124_integer_4`, \
                zbld_125_integer_5, `zbld_131_decimal_1`, `zbld_132_decimal_2`, \
                zbld_133_decimal_3, `zbld_134_decimal_4`, `zbld_135_decimal_5`, \
                zbld_141_date_1, `zbld_142_date_2`, `zbld_143_date_3`, \
                zbld_144_date_4, `zbld_145_date_5`, `v_client_pk_consigment_num`, \
                z_createdTimeStamp, `z_createdByAccount`, `z_modifiedTimeStamp`, \
                `z_modifiedByAccount`, `success`,  `fk_booking_lines_id`) \
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, \
                  %s, %s, %s, %s, %s, %s)"
    args = [v for _, v in line_detail.items()]
    dbcur.execute(sql, args)


def insert_log(dbcon, fname, success, date):
    sql = "INSERT INTO Log (__id, filename, success, date) VALUES (0, %s, %s, %s)"
    args = [fname, "1" if success else "0", date.isoformat()]
    with dbcon.cursor() as cur:
        cur.execute(sql, args)


def checkLen(str, limit):
    if str is not None:
        if len(str) > limit:
            return False

    return True


def _strip(value):
    if value and isinstance(value, str):
        return value.strip()
    else:
        return value


def do_import(dbcon, cur, filename):
    wb = load_workbook(filename=filename, data_only=True)
    warehouse_id = []
    warehouse_name = []
    warehouse_code = []
    warehouse_success_type = []
    import_type = 0

    if (
        "Import Template Headers" in wb.sheetnames
        and "Import Template Lines" in wb.sheetnames
        and "Related Line Data Import" in wb.sheetnames
    ):
        worksheet0 = wb["Import Template Headers"]
        worksheet1 = wb["Import Template Lines"]
        worksheet2 = wb["Related Line Data Import"]
        print("Import Template Headers")
    else:
        return

    warehouse_code.append("No - Warehouse")
    warehouse_name.append("No - Warehouse")
    warehouse_id.append(100)
    warehouse_success_type.append("2")

    # Header
    row = 4
    headers = []

    while True:
        if worksheet0["A%i" % row].value == None:
            break

        header = {}
        print(f"#801 - Reading HEADER(S): {row - 3}")
        pk_header_id = str(uuid.uuid1())
        header["id"] = worksheet0["A%i" % row].value
        header["client_booking_id"] = pk_header_id
        header["b_003_b_service_name"] = "R"
        header["client_cust_job_code"] = None
        header["company"] = _strip(worksheet0["AI%i" % row].value)
        header["total_lines"] = None
        header["address_suburb"] = _strip(worksheet0["AM%i" % row].value)
        header["address_state"] = _strip(worksheet0["AN%i" % row].value)
        header["address_postalcode"] = worksheet0["AP%i" % row].value
        header["v_client_pk_consigment_num"] = pk_header_id
        header["total_kg"] = None
        header["fk_client_warehouse_id"] = warehouse_id[0]
        header["puPickUpAvailFrom_Date"] = worksheet0["C%i" % row].value
        header["b_022_b_pu_avail_from_time_hour"] = worksheet0["D%i" % row].value
        header["b_024_b_pu_by_date"] = worksheet0["F%i" % row].value
        header["b_025_b_pu_by_time_hour"] = worksheet0["G%i" % row].value
        header["b_047_b_del_avail_from_date"] = worksheet0["AC%i" % row].value
        header["b_048_b_del_avail_from_time_hour"] = worksheet0["AD%i" % row].value
        header["b_050_b_del_by_date"] = worksheet0["AF%i" % row].value
        header["b_051_b_del_by_time_hour"] = worksheet0["AG%i" % row].value
        header["b_client_warehouse_code"] = _strip(warehouse_code[0])
        header["b_028_b_pu_company"] = _strip(worksheet0["I%i" % row].value)
        header["b_029_b_pu_address_street_1"] = _strip(worksheet0["K%i" % row].value)
        header["b_030_b_pu_address_street_2"] = _strip(worksheet0["L%i" % row].value)
        header["b_031_b_pu_address_state"] = _strip(worksheet0["N%i" % row].value)
        header["b_033_b_pu_address_postalcode"] = worksheet0["P%i" % row].value
        header["b_032_b_pu_address_suburb"] = _strip(worksheet0["M%i" % row].value)
        header["b_035_b_pu_contact_full_name"] = _strip(worksheet0["R%i" % row].value)
        header["b_038_b_pu_phone_main"] = worksheet0["S%i" % row].value
        header["b_037_b_pu_email"] = _strip(worksheet0["U%i" % row].value)
        header["b_015_b_pu_instructions_contact"] = _strip(
            worksheet0["Z%i" % row].value
        )
        header["b_055_b_del_address_street_1"] = _strip(worksheet0["AK%i" % row].value)
        header["b_044_b_del_instructions_address"] = _strip(
            worksheet0["AY%i" % row].value
        )
        header["b_043_b_del_instructions_contact"] = _strip(
            worksheet0["AX%i" % row].value
        )
        header["b_064_b_del_phone_main"] = worksheet0["AS%i" % row].value
        header["b_063_b_del_email"] = _strip(worksheet0["AU%i" % row].value)
        header["b_005_b_created_for"] = _strip(worksheet0["BG%i" % row].value)
        header["b_006_b_created_for_email"] = None
        header["b_000_1_b_clientReference_RA_Numbers"] = _strip(
            worksheet0["BC%i" % row].value
        )
        header["z_test"] = None
        header["b_client_sales_inv_num"] = None
        header["b_client_order_num"] = None
        header["b_client_del_note_num"] = None
        header["date_processed"] = convert_to_UTC_tz(datetime.now())
        header["z_createdTimeStamp"] = convert_to_UTC_tz(datetime.now())
        header["success"] = warehouse_success_type[0]
        header["b_007_b_ready_status"] = worksheet0["B%i" % row].value
        header["b_023_b_pu_avail_from_time_minute"] = worksheet0["E%i" % row].value
        header["b_026_b_pu_by_time_minute"] = worksheet0["H%i" % row].value
        header["b_clientPU_Warehouse"] = warehouse_name[0]
        header["b_027_b_pu_address_type"] = worksheet0["J%i" % row].value
        header["b_034_b_pu_address_country"] = _strip(worksheet0["O%i" % row].value)
        header["pu_addressed_saved"] = worksheet0["Q%i" % row].value
        header["b_039_b_pu_phone_mobile"] = worksheet0["T%i" % row].value
        header["b_036_b_pu_email_group"] = _strip(worksheet0["V%i" % row].value)
        header["b_040_b_pu_communicate_via"] = worksheet0["W%i" % row].value
        header["b_014_b_pu_handling_instructions"] = worksheet0["X%i" % row].value
        header["b_016_b_pu_instructions_address"] = worksheet0["Y%i" % row].value
        header["b_017_b_pu_warehouse_num"] = worksheet0["AA%i" % row].value
        header["b_018_b_pu_warehouse_bay"] = worksheet0["AB%i" % row].value
        header["b_049_b_del_avail_from_time_minute"] = worksheet0["AE%i" % row].value
        header["b_052_b_del_by_time_minute"] = worksheet0["AH%i" % row].value
        header["b_053_b_del_address_type"] = worksheet0["AJ%i" % row].value
        header["b_056_b_del_address_street_2"] = _strip(worksheet0["AL%i" % row].value)
        header["b_060_b_del_address_country"] = _strip(worksheet0["AO%i" % row].value)
        header["de_to_addressed_saved"] = worksheet0["AQ%i" % row].value
        header["b_061_b_del_contact_full_name"] = _strip(worksheet0["AR%i" % row].value)
        header["b_065_b_del_phone_mobile"] = worksheet0["AT%i" % row].value
        header["b_062_b_del_email_group"] = _strip(worksheet0["AV%i" % row].value)
        header["b_066_b_del_communicate_via"] = worksheet0["AW%i" % row].value
        header["b_046_b_del_warehouse_number"] = worksheet0["AZ%i" % row].value
        header["b_045_b_del_warehouse_bay"] = worksheet0["BA%i" % row].value
        header["b_010_b_notes"] = _strip(worksheet0["BB%i" % row].value)
        header["b_008_b_category"] = _strip(worksheet0["BD%i" % row].value)
        header["b_client_max_book_amount"] = worksheet0["BE%i" % row].value
        header["b_100_client_price_paid_or_quoted"] = None
        header["b_009_b_priority"] = worksheet0["BF%i" % row].value

        if (
            not worksheet0["BH%i" % row].value == None
            and str(worksheet0["BH%i" % row].value).lower() == "no"
        ):
            header["b_012_b_driver_bring_connote"] = 0
        else:
            header["b_012_b_driver_bring_connote"] = 1

        if (
            not worksheet0["BI%i" % row].value == None
            and str(worksheet0["BI%i" % row].value).lower() == "no"
        ):
            header["b_019_b_pu_tail_lift"] = 0
        else:
            header["b_019_b_pu_tail_lift"] = 1

        header["b_020_b_pu_num_operators"] = worksheet0["BJ%i" % row].value

        if (
            not worksheet0["BK%i" % row].value == None
            and str(worksheet0["BK%i" % row].value).lower() == "no"
        ):
            header["b_041_b_del_tail_lift"] = 0
        else:
            header["b_041_b_del_tail_lift"] = 1

        header["b_042_b_del_num_operators"] = worksheet0["BL%i" % row].value

        if (
            not worksheet0["BM%i" % row].value == None
            and str(worksheet0["BM%i" % row].value).lower() == "no"
        ):
            header["b_013_b_package_job"] = 0
        else:
            header["b_013_b_package_job"] = 1

        header["b_001_b_freight_provider"] = None
        header["vx_serviceType_XXX"] = None
        header["fp_pu_id"] = None
        header["b_002_b_vehicle_type"] = None
        header["zb_101_text_1"] = None
        header["zb_102_text_2"] = None
        header["zb_103_text_3"] = None
        header["zb_104_text_4"] = None
        header["zb_105_Text_5"] = None
        header["zb_121_integer_1"] = None
        header["zb_122_integer_2"] = None
        header["zb_123_integer_3"] = None
        header["zb_124_integer_4"] = None
        header["zb_125_integer_5"] = None
        header["zb_131_decimal_1"] = None
        header["zb_132_decimal_2"] = None
        header["zb_133_decimal_3"] = None
        header["zb_134_decimal_4"] = None
        header["zb_135_decimal_5"] = None
        header["zb_141_date_1"] = None
        header["zb_142_date_2"] = None
        header["zb_143_date_3"] = None
        header["zb_144_date_4"] = None
        header["zb_145_date_5"] = None
        header["pk_header_id"] = pk_header_id
        header["fk_client_id"] = "37C19636-C5F9-424D-AD17-05A056A8FBDB"
        header["b_000_3_consignment_number"] = ""
        header["x_booking_Created_With"] = "Imported(Sharepoint)"

        headers.append(header)
        row += 1

    # Line
    row = 3
    lines = []

    while True:
        if worksheet1["A%i" % row].value == None:
            break

        line = {}
        print(f"#802 - Reading LINE: {row - 2}")
        line["id"] = worksheet1["A%i" % row].value
        line["header_id"] = worksheet1["B%i" % row].value
        line["client_booking_id"] = None
        line["client_UOM"] = None
        line["weight_per_each"] = worksheet1["G%i" % row].value
        line["l_010_totaldim"] = None
        line["client_run_code"] = None
        line["l_item"] = _strip(worksheet1["D%i" % row].value)
        line["v_client_pk_consigment_num"] = None
        line["l_cubic_weight"] = None
        line["qty"] = worksheet1["E%i" % row].value
        line["l_001_type_of_packaging"] = worksheet1["C%i" % row].value
        line["l_005_dim_length"] = worksheet1["I%i" % row].value
        line["l_006_dim_width"] = worksheet1["J%i" % row].value
        line["l_007_dim_height"] = worksheet1["K%i" % row].value
        line["date_processed"] = convert_to_UTC_tz(datetime.now())
        line["z_createdTimeStamp"] = convert_to_UTC_tz(datetime.now())
        line["success"] = warehouse_success_type[0]
        line["e_pallet_type"] = None
        line["client_item_number"] = None
        line["e_item_type"] = None
        line["client_item_reference"] = None
        line["l_004_dim_UOM"] = worksheet1["H%i" % row].value
        line["l_008_weight_UOM"] = worksheet1["F%i" % row].value
        line["zbl_101_text_1"] = None
        line["zbl_102_text_2"] = None
        line["zbl_103_text_3"] = None
        line["zbl_104_text_4"] = None
        line["zbl_105_Text_5"] = None
        line["zbl_121_integer_1"] = None
        line["zbl_122_integer_2"] = None
        line["zbl_123_integer_3"] = None
        line["zbl_124_integer_4"] = None
        line["zbl_125_integer_5"] = None
        line["zbl_131_decimal_1"] = None
        line["zbl_132_decimal_2"] = None
        line["zbl_133_decimal_3"] = None
        line["zbl_134_decimal_4"] = None
        line["zbl_135_decimal_5"] = None
        line["zbl_141_date_1"] = None
        line["zbl_142_date_2"] = None
        line["zbl_143_date_3"] = None
        line["zbl_144_date_4"] = None
        line["zbl_145_date_5"] = None
        line["pk_booking_lines_id"] = str(uuid.uuid1())

        lines.append(line)
        row += 1

    # Line details
    row = 3
    line_details = []

    while True:
        if worksheet2["A%i" % row].value == None:
            break

        print(f"#803 - Reading LINE DETAIL: {row - 2}")
        line_detail = {}
        line_detail["id"] = worksheet2["A%i" % row].value
        line_detail["line_id"] = worksheet2["B%i" % row].value
        line_detail["client_booking_id"] = None
        line_detail["ld_002_model_number"] = _strip(worksheet2["C%i" % row].value)
        line_detail["ld_003_item_description"] = _strip(worksheet2["D%i" % row].value)
        line_detail["ld_001_qty"] = worksheet2["E%i" % row].value
        line_detail["ld_004_fault_description"] = _strip(worksheet2["G%i" % row].value)
        line_detail["ld_007_gap_ra"] = _strip(worksheet2["H%i" % row].value)
        line_detail["ld_005_item_serial_number"] = _strip(worksheet2["I%i" % row].value)
        line_detail["ld_006_insurance_value"] = _strip(worksheet2["J%i" % row].value)
        line_detail["ld_008_client_ref_number"] = _strip(worksheet2["K%i" % row].value)

        line_detail["zbld_101_text_1"] = ""
        line_detail["zbld_102_text_2"] = ""
        line_detail["zbld_103_text_3"] = ""
        line_detail["zbld_104_text_4"] = ""
        line_detail["zbld_105_Text_5"] = ""
        line_detail["zbld_121_integer_1"] = None
        line_detail["zbld_122_integer_2"] = None
        line_detail["zbld_123_integer_3"] = None
        line_detail["zbld_124_integer_4"] = None
        line_detail["zbld_125_integer_5"] = None
        line_detail["zbld_131_decimal_1"] = None
        line_detail["zbld_132_decimal_2"] = None
        line_detail["zbld_133_decimal_3"] = None
        line_detail["zbld_134_decimal_4"] = None
        line_detail["zbld_135_decimal_5"] = None
        line_detail["zbld_141_date_1"] = convert_to_UTC_tz(datetime.now())
        line_detail["zbld_142_date_2"] = convert_to_UTC_tz(datetime.now())
        line_detail["zbld_143_date_3"] = convert_to_UTC_tz(datetime.now())
        line_detail["zbld_144_date_4"] = convert_to_UTC_tz(datetime.now())
        line_detail["zbld_145_date_5"] = convert_to_UTC_tz(datetime.now())

        line_detail["v_client_pk_consigment_num"] = None
        line_detail["z_createdTimeStamp"] = convert_to_UTC_tz(datetime.now())
        line_detail["z_createdByAccount"] = ""
        line_detail["z_modifiedTimeStamp"] = convert_to_UTC_tz(datetime.now())
        line_detail["z_modifiedByAccount"] = ""
        line_detail["success"] = warehouse_success_type[0]

        line_details.append(line_detail)
        row += 1

    # Update File note
    pk_header_ids = []
    for header in headers:
        pk_header_ids.append(header["pk_header_id"])

    print(f"#805 - Updating files table note info ...", cur.lastrowid)
    save_dme_note(dbcon, cur.lastrowid, ", ".join(pk_header_ids))

    # Insert to DB
    print(f"#804 - Inserting to DB...")
    pk_header_ids = []
    for header in headers:
        pk_header_ids.append(header["pk_header_id"])

        for line in lines:
            if line["header_id"] == header["id"]:
                line["client_booking_id"] = header["pk_header_id"]
                line["v_client_pk_consigment_num"] = header["pk_header_id"]

                for line_detail in line_details:
                    if line_detail["line_id"] == line["id"]:
                        line_detail["client_booking_id"] = header["pk_header_id"]
                        line_detail["v_client_pk_consigment_num"] = header[
                            "pk_header_id"
                        ]
                        line_detail["fk_booking_lines_id"] = line["pk_booking_lines_id"]

                        insert_line_detail(cur, delete_keys(line_detail))
                insert_line(cur, delete_keys(line))
        insert_header(cur, delete_keys(header))

    dbcon.commit()
    return True


def send_error_email(filename, error):
    text = f"There is an issue with this file ({filename}) \n\n error is {error} \n\nplease correct the file and rename it with _x2, x3 etc and re-drop it into the drop folder."
    send_email(
        ["bookings@deliver-me.com.au"],
        ["dev.deliverme@gmail.com", "goldj@deliver-me.com.au"],
        "Error",
        text,
    )


def send_duplicate_email(filename):
    text = f"file ({filename}) (is a duplicate file) \n\nPlease check the file in error directory and dme system. \n\nIf it needs to be imported, change   the file name and re-drop it into the drop folder."
    send_email(
        ["bookings@deliver-me.com.au"],
        ["dev.deliverme@gmail.com", "goldj@deliver-me.com.au"],
        "Duplicated",
        text,
    )


def download_from_sharepoint(dbcon):
    cur = dbcon.cursor()

    share_point_site = "https://delivermeee.sharepoint.com"
    username = "goldj@deliver-me.com.au"
    password = "qos131QOS131*"
    site_url = "https://delivermeee.sharepoint.com/sites/DeliverMECustomerSite/"
    folderpath = "Shared Documents/Bookings/Tempo/1_Tempo Drop"
    completed_path = (
        f"/sites/DeliverMECustomerSite/Shared Documents/Bookings/Tempo/3_Completed"
    )
    error_path = (
        f"/sites/DeliverMECustomerSite/Shared Documents/Bookings/Tempo/4_ErrorToCheck"
    )
    DEV_SUFFIX = "_Test"

    if DB_NAME in ["dme_db_dev", "deliver_me"]:
        folderpath += DEV_SUFFIX
        completed_path += DEV_SUFFIX
        error_path += DEV_SUFFIX

    authcookie = Office365(
        share_point_site=share_point_site, username=username, password=password
    ).get_cookies()

    site = Site(site_url=site_url, authcookie=authcookie)

    archive_dir = IMPORT_DIR
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)

    if not os.path.exists(ERROR_DIR):
        os.makedirs(ERROR_DIR)

    archive_files = os.listdir(archive_dir)

    files = site.get_files(folderpath=folderpath)

    for file in files:
        filename = file["Name"]
        filepath = file["ServerRelativeUrl"]
        print("File path on sharepoint: ", filepath)

        if not ".xls" in filename.lower():
            print(f"File format is incorrect: {filename}")
        elif filename in archive_files:
            print(f"File is already imported: {filename}")
            send_duplicate_email(filename)
        elif ".xls" in filename.lower() and not filename in archive_files:
            print("#601 File downloading... ", filepath)
            file_content = site.download_file(
                filename=filename, filepath=filepath, downloadpath=DOWNLOAD_DIR
            )
            print("#602 File downloaded!")

            dme_file = {}
            dme_file["file_name"] = filename
            dme_file["z_createdTimeStamp"] = convert_to_UTC_tz(datetime.now())
            dme_file["z_createdByAccount"] = "DME"
            dme_file["file_type"] = "xls import"
            dme_file["file_path"] = str(
                os.path.abspath(os.path.join(DOWNLOAD_DIR, filename))
            )
            dme_file["file_extension"] = "xlsx"

            insert_dme_file(cur, dme_file)
            dbcon.commit()

            for filename in os.listdir(DOWNLOAD_DIR):
                fpath = os.path.join(DOWNLOAD_DIR, filename)

                if os.path.isfile(fpath) and filename.lower().endswith(".xlsx"):
                    try:
                        pk_header_id = do_import(dbcon, cur, fpath)
                        shutil.move(fpath, os.path.join(IMPORT_DIR, filename))
                        save_dme_file_path(
                            dbcon,
                            cur.lastrowid,
                            str(os.path.abspath(os.path.join(IMPORT_DIR, filename))),
                        )
                        path = (
                            completed_path + "/" + datetime.now().strftime("%Y-%m-%d")
                        )
                        site.create_folder(folderpath=path)
                        site.move_file(
                            src_relative_path=filepath, dest_path=path + "/" + filename
                        )
                    except OSError as e:
                        print("#1020 Error", str(e))

                        save_dme_note(dbcon, cur.lastrowid, str(e))
                        shutil.move(fpath, os.path.join(ERROR_DIR, filename))
                        save_dme_file_path(
                            dbcon,
                            cur.lastrowid,
                            str(os.path.abspath(os.path.join(ERROR_DIR, filename))),
                        )
                        path = error_path + "/" + datetime.now().strftime("%Y-%m-%d")
                        site.create_folder(folderpath=path)
                        site.move_file(
                            src_relative_path=filepath, dest_path=path + "/" + filename
                        )

                        send_error_email(filename, str(e))
                        success = False
                    except Exception as e:
                        print("#1030 Error", str(e))
                        traceback.print_exc()

                        save_dme_note(dbcon, cur.lastrowid, str(e))
                        shutil.move(fpath, os.path.join(ERROR_DIR, filename))
                        save_dme_file_path(
                            dbcon,
                            cur.lastrowid,
                            str(os.path.abspath(os.path.join(ERROR_DIR, filename))),
                        )
                        path = error_path + "/" + datetime.now().strftime("%Y-%m-%d")
                        site.create_folder(folderpath=path)
                        site.move_file(
                            src_relative_path=filepath, dest_path=path + "/" + filename
                        )
                        success = False
                        send_error_email(filename, str(e))
                    # insert_log(dbcon, fname, success, datetime.now())
            dbcon.commit()


if __name__ == "__main__":
    print("Started %s" % datetime.now())
    time1 = time.time()

    try:
        dbcon = pymysql.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASS,
            db=DB_NAME,
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
        )
    except:
        print("Mysql DB connection error!")
        exit(1)

    try:
        option = get_option(dbcon, "xls_import")

        if int(option["option_value"]) == 0:
            print("#905 - `xls_import` option is OFF")
        elif option["is_running"]:
            print("#905 - `xls_import` script is already RUNNING")
        else:
            print("#906 - `xls_import` option is ON")
            set_option(dbcon, "xls_import", True)
            print("#910 - Processing...")
            download_from_sharepoint(dbcon)
            set_option(dbcon, "xls_import", False, time1)
    except Exception as e:
        print("Error 904:", str(e))
        set_option(dbcon, "xls_import", False, time1)

    dbcon.close()
    print("#999 - Finished %s" % datetime.now())
